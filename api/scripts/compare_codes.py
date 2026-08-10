from pathlib import Path
from openpyxl import load_workbook
import psycopg2
import os


def load_env():
    env_path = Path(__file__).parent.parent / ".env"
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, value = line.split("=", 1)
                    os.environ.setdefault(key.strip(), value.strip())


def main():
    load_env()

    # Leer códigos del Excel
    excel_path = Path("Y:/ALMACEN/Mejora Continua ALMACEN/Nuevo Control de Almacen/BASE DE DATOS/BD_ALMACEN_3P.xlsx")
    wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    ws = wb["DETALLE_VALES"]
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    excel_codes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        item = dict(zip(headers, row))
        codigo = item.get("CODIGO")
        if codigo is not None:
            excel_codes.add(str(codigo).strip())
    wb.close()

    # Leer códigos de sae_existencias
    conn = psycopg2.connect(
        host=os.environ.get("POSTGRES_HOST", "localhost"),
        port=os.environ.get("POSTGRES_PORT", "5432"),
        dbname=os.environ.get("POSTGRES_DB", "cj_assistant"),
        user=os.environ.get("POSTGRES_USER", ""),
        password=os.environ.get("POSTGRES_PASSWORD", ""),
    )
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT cve_art FROM sae_existencias WHERE exist > 0;")
    sae_codes = {str(r[0]).strip() for r in cur.fetchall()}
    conn.close()

    print(f"Códigos en Excel (DETALLE_VALES): {len(excel_codes)}")
    print(f"Códigos en sae_existencias: {len(sae_codes)}")
    print(f"Coincidencias exactas: {len(excel_codes & sae_codes)}")
    print(f"\nPrimeros 10 códigos del Excel:")
    for c in sorted(excel_codes)[:10]:
        print(f"  {c}")
    print(f"\nPrimeros 10 códigos de SAE:")
    for c in sorted(sae_codes)[:10]:
        print(f"  {c}")


if __name__ == "__main__":
    main()
