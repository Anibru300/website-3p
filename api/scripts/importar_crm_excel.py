#!/usr/bin/env python3
"""
Importa datos de CRM desde los archivos Excel de Y:/CRM´S/ hacia SQLite.

Uso:
    .venv/Scripts/python.exe scripts/importar_crm_excel.py
    .venv/Scripts/python.exe scripts/importar_crm_excel.py --dir "Y:/CRM'S"
    .venv/Scripts/python.exe scripts/importar_crm_excel.py --actualizar

REGLA DE VERDAD ÚNICA (Fase 3): el panel admin es la fuente de verdad del
CRM. Por defecto este script opera en MODO SOLO-ALTAS:
- Inserta entidades/registros que no existan en SQLite.
- NO actualiza registros que ya existan (respeta las ediciones hechas en
  el panel admin) y evita duplicar INSERTs al re-correrlo.
Solo con el flag --actualizar recupera el comportamiento anterior de
pisar los datos existentes con lo que traiga el Excel.

El script:
- Lee CRM´S FINAL 3P.xlsm (hojas CAT_CLIENTES, CAT_GRANJAS, CAT_DOMICILIOS, CAT_PAQUETERIAS)
- Lee CRM PORTALES CLIENTES SUBIR FACTURAS.xlsx
- Lee CRM CLIENTES.xlsx (contactos)
- Lee CRM DESCUENTOS CLIENTES.xlsx (días de crédito y descuentos)
- Hace matching por nombre de cliente normalizado
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl no está instalado. Instálalo con: pip install openpyxl"
    ) from exc

from app.database import users_connection

DEFAULT_DIR = Path("Y:/CRM´S")
ALT_DIR = Path("Y:/CRM'S")


def normalizar_nombre(texto):
    """Normaliza un nombre para matching: minúsculas, sin espacios extra ni puntuación."""
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = re.sub(r"\s+", " ", texto)
    texto = re.sub(r"[^a-z0-9áéíóúüñ ]", "", texto)
    return texto.strip()


def limpiar_valor(valor):
    """Convierte celdas vacías en None y limpia espacios."""
    if valor is None:
        return None
    if isinstance(valor, str):
        valor = valor.strip()
        if valor == "":
            return None
        return valor
    return valor


def obtener_columnas(ws):
    """Devuelve un diccionario {nombre_normalizado: letra_columna} desde la primera fila."""
    headers = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[normalizar_nombre(str(cell.value))] = cell.column_letter
    return headers


def valor_celda(ws, fila, letra):
    if letra is None:
        return None
    return limpiar_valor(ws[f"{letra}{fila}"].value)


def encontrar_o_crear_entidad(conn, nombre, defaults=None):
    """Busca entidad por nombre normalizado. Si no existe, la crea con defaults.

    Devuelve (id, creada): creada=True si se insertó en esta corrida.
    """
    nombre_limpio = limpiar_valor(nombre)
    if not nombre_limpio:
        return None, False
    nombre_str = str(nombre_limpio)
    norm = normalizar_nombre(nombre_str)

    row = conn.execute(
        "SELECT id FROM crm_entidades WHERE activo = 1 AND lower(trim(nombre)) = ?",
        (nombre_str.lower(),),
    ).fetchone()

    if row:
        return row["id"], False

    # Intento con normalización flexible
    cursor = conn.execute(
        "SELECT id, nombre FROM crm_entidades WHERE activo = 1"
    )
    for r in cursor:
        if normalizar_nombre(r["nombre"]) == norm:
            return r["id"], False

    defaults = defaults or {}
    now = datetime.utcnow().isoformat()
    cur = conn.execute(
        """
        INSERT INTO crm_entidades (
            id_externo, tipo, nombre, razon_social, rfc, telefono, email,
            condicion_pago, dias_credito, vendedor, industria, status, notas, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            defaults.get("id_externo"),
            defaults.get("tipo", "cliente"),
            nombre_str,
            defaults.get("razon_social"),
            defaults.get("rfc"),
            defaults.get("telefono"),
            defaults.get("email"),
            defaults.get("condicion_pago"),
            defaults.get("dias_credito"),
            defaults.get("vendedor"),
            defaults.get("industria"),
            defaults.get("status", "Activo"),
            defaults.get("notas"),
            now,
        ),
    )
    conn.commit()
    return cur.lastrowid, True


def buscar_entidad_id(conn, nombre):
    """Solo busca entidad existente, sin crear."""
    nombre_limpio = limpiar_valor(nombre)
    if not nombre_limpio:
        return None
    nombre_str = str(nombre_limpio)
    norm = normalizar_nombre(nombre_str)

    row = conn.execute(
        "SELECT id FROM crm_entidades WHERE activo = 1 AND lower(trim(nombre)) = ?",
        (nombre_str.lower(),),
    ).fetchone()
    if row:
        return row["id"]

    cursor = conn.execute("SELECT id, nombre FROM crm_entidades WHERE activo = 1")
    for r in cursor:
        if normalizar_nombre(r["nombre"]) == norm:
            return r["id"]
    return None


def buscar_entidad_por_id_externo(conn, id_externo):
    """Busca entidad por id_externo."""
    if id_externo is None:
        return None
    row = conn.execute(
        "SELECT id FROM crm_entidades WHERE activo = 1 AND id_externo = ?",
        (str(id_externo).strip(),),
    ).fetchone()
    return row["id"] if row else None


def buscar_entidad_flexible(conn, nombre):
    """Búsqueda flexible por nombre: id_externo, exacto, normalizado, substring."""
    if not nombre:
        return None

    # Por id_externo si es numérico
    entidad_id = buscar_entidad_por_id_externo(conn, nombre)
    if entidad_id:
        return entidad_id

    # Exacto y normalizado
    entidad_id = buscar_entidad_id(conn, nombre)
    if entidad_id:
        return entidad_id

    # Búsqueda por substring
    nombre_str = str(limpiar_valor(nombre)).lower()
    norm = normalizar_nombre(nombre_str)

    cursor = conn.execute("SELECT id, nombre FROM crm_entidades WHERE activo = 1")
    mejores = []
    for r in cursor:
        nombre_db = str(r["nombre"]).lower()
        norm_db = normalizar_nombre(r["nombre"])
        if nombre_str in nombre_db or nombre_db in nombre_str:
            mejores.append((r["id"], len(nombre_db)))
        elif norm and (norm in norm_db or norm_db in norm):
            mejores.append((r["id"], len(norm_db)))

    if mejores:
        # Elegir el más corto (más específico)
        mejores.sort(key=lambda x: x[1])
        return mejores[0][0]

    return None


def mapear_columnas(headers, mapeos):
    """
    mapeos: lista de tuplas (clave_salida, [nombres_candidatos_normalizados])
    Devuelve dict {clave_salida: letra_columna}
    """
    resultado = {}
    for clave, candidatos in mapeos:
        for c in candidatos:
            if c in headers:
                resultado[clave] = headers[c]
                break
    return resultado


# ---------------------------------------------------------------------------
# Procesadores por hoja/archivo
# ---------------------------------------------------------------------------


def procesar_cat_clientes(conn, ruta, actualizar=False):
    """Lee CAT_CLIENTES y crea entidades nuevas.

    Modo solo-altas (default): las entidades que ya existen no se tocan
    (respeta ediciones del panel admin). Con actualizar=True recupera el
    comportamiento anterior de pisar datos.
    """
    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb["CAT_CLIENTES"]
    headers = obtener_columnas(ws)

    mapeos = [
        ("id_externo", ["clienteid", "id cliente", "cliente id", "id"]),
        ("nombre", ["nombre cliente", "nombre", "cliente", "razon social", "razón social"]),
        ("razon_social", ["razon social", "razón social", "nombre fiscal"]),
        ("rfc", ["rfc"]),
        ("telefono", ["telefono", "teléfono", "tel"]),
        ("email", ["correo cfdi", "correo", "email", "e-mail", "correo electronico", "correo electrónico"]),
        ("dias_credito", ["dias credito", "días crédito", "dias de credito", "credito", "crédito"]),
        ("vendedor", ["vendedor", "asesor", "ejecutivo"]),
        ("industria", ["industria", "giro", "sector"]),
        ("status", ["status", "estatus", "estado"]),
        ("notas", ["notas", "observaciones", "comentarios"]),
    ]
    cols = mapear_columnas(headers, mapeos)

    creadas = 0
    actualizadas = 0
    omitidas = 0

    for fila in range(2, ws.max_row + 1):
        nombre = valor_celda(ws, fila, cols.get("nombre", "B"))
        if not nombre:
            omitidas += 1
            continue

        id_externo = valor_celda(ws, fila, cols.get("id_externo", "A"))

        defaults = {
            "id_externo": str(id_externo) if id_externo is not None else None,
            "tipo": "cliente",
            "razon_social": valor_celda(ws, fila, cols.get("razon_social")) or nombre,
            "rfc": valor_celda(ws, fila, cols.get("rfc")),
            "telefono": valor_celda(ws, fila, cols.get("telefono")),
            "email": valor_celda(ws, fila, cols.get("email")),
            "dias_credito": str(valor_celda(ws, fila, cols.get("dias_credito"))) if cols.get("dias_credito") else None,
            "vendedor": valor_celda(ws, fila, cols.get("vendedor")),
            "industria": valor_celda(ws, fila, cols.get("industria")),
            "status": valor_celda(ws, fila, cols.get("status")) or "Activo",
            "notas": valor_celda(ws, fila, cols.get("notas")),
        }

        entidad_id, creada = encontrar_o_crear_entidad(conn, nombre, defaults)
        if entidad_id:
            if creada:
                creadas += 1
            elif not actualizar:
                omitidas += 1  # Ya existía: no se pisan ediciones del panel
            else:
                # Actualizar campos si la entidad ya existía
                campos = {k: v for k, v in defaults.items() if k != "tipo" and v is not None}
                if campos:
                    campos["updated_at"] = datetime.utcnow().isoformat()
                    set_clause = ", ".join(f"{k} = ?" for k in campos.keys())
                    conn.execute(
                        f"UPDATE crm_entidades SET {set_clause} WHERE id = ?",
                        (*campos.values(), entidad_id),
                    )
                    conn.commit()
                    actualizadas += 1
                else:
                    omitidas += 1

    wb.close()
    return {"creadas": creadas, "actualizadas": actualizadas, "omitidas": omitidas}


def procesar_cat_granjas(conn, ruta, actualizar=False):
    """Lee CAT_GRANJAS. Solo-altas: no actualiza granjas existentes."""
    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb["CAT_GRANJAS"]
    headers = obtener_columnas(ws)

    mapeos = [
        ("cliente_id", ["clienteid", "cliente id", "id cliente", "cliente"]),
        ("granja_id", ["id granja", "granja id", "id", "clave granja"]),
        ("nombre", ["granja", "nombre granja", "nombre de granja"]),
        ("tipo", ["tipo", "tipo granja"]),
        ("paso", ["paso", "fase", "etapa"]),
        ("contacto", ["contacto", "nombre contacto", "contacto granja"]),
        ("puesto", ["puesto", "puesto contacto"]),
        ("telefono", ["telefono", "teléfono", "tel"]),
        ("correo", ["correo", "email", "correo contacto"]),
        ("comentarios", ["comentarios", "observaciones", "notas"]),
    ]
    cols = mapear_columnas(headers, mapeos)

    creadas = 0
    actualizadas = 0
    omitidas = 0

    for fila in range(2, ws.max_row + 1):
        cliente_id = valor_celda(ws, fila, cols.get("cliente_id", "A"))
        nombre = valor_celda(ws, fila, cols.get("nombre", "B"))
        if not cliente_id or not nombre:
            omitidas += 1
            continue

        # Buscar por id_externo primero
        entidad_id = buscar_entidad_por_id_externo(conn, cliente_id)
        if not entidad_id:
            entidad_id = buscar_entidad_id(conn, cliente_id)

        if not entidad_id:
            print(f"  [granjas] No se encontró entidad para ClienteID '{cliente_id}' (granja '{nombre}'). Omitiendo.")
            omitidas += 1
            continue

        # Evitar duplicados por entidad + nombre de granja
        existente = conn.execute(
            "SELECT id FROM crm_granjas WHERE entidad_id = ? AND lower(trim(nombre)) = ? AND activo = 1",
            (entidad_id, nombre.lower()),
        ).fetchone()

        if existente:
            if not actualizar:
                omitidas += 1  # Ya existía: no se pisan ediciones del panel
                continue
            conn.execute(
                """
                UPDATE crm_granjas SET
                    granja_id_externo = ?, tipo = ?, paso = ?, contacto_nombre = ?,
                    contacto_puesto = ?, contacto_telefono = ?, contacto_correo = ?, comentarios = ?
                WHERE id = ?
                """,
                (
                    valor_celda(ws, fila, cols.get("granja_id")),
                    valor_celda(ws, fila, cols.get("tipo")),
                    valor_celda(ws, fila, cols.get("paso")),
                    valor_celda(ws, fila, cols.get("contacto")),
                    valor_celda(ws, fila, cols.get("puesto")),
                    valor_celda(ws, fila, cols.get("telefono")),
                    valor_celda(ws, fila, cols.get("correo")),
                    valor_celda(ws, fila, cols.get("comentarios")),
                    existente["id"],
                ),
            )
            actualizadas += 1
        else:
            conn.execute(
                """
                INSERT INTO crm_granjas (
                    entidad_id, granja_id_externo, nombre, tipo, paso,
                    contacto_nombre, contacto_puesto, contacto_telefono, contacto_correo, comentarios
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    entidad_id,
                    valor_celda(ws, fila, cols.get("granja_id")),
                    nombre,
                    valor_celda(ws, fila, cols.get("tipo")),
                    valor_celda(ws, fila, cols.get("paso")),
                    valor_celda(ws, fila, cols.get("contacto")),
                    valor_celda(ws, fila, cols.get("puesto")),
                    valor_celda(ws, fila, cols.get("telefono")),
                    valor_celda(ws, fila, cols.get("correo")),
                    valor_celda(ws, fila, cols.get("comentarios")),
                ),
            )
        conn.commit()
        creadas += 1

    wb.close()
    return {"creadas": creadas, "actualizadas": actualizadas, "omitidas": omitidas}


def procesar_cat_domicilios(conn, ruta, actualizar=False):
    """Lee CAT_DOMICILIOS. Solo-altas: evita duplicar domicilios al re-correr."""
    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb["CAT_DOMICILIOS"]
    headers = obtener_columnas(ws)

    mapeos = [
        ("cliente", ["cliente", "nombre cliente", "empresa"]),
        ("granja", ["granja", "nombre granja"]),
        ("nombre", ["nombre", "alias", "tipo", "tipo domicilio"]),
        ("calle", ["calle"]),
        ("numero", ["numero", "número", "numero exterior", "no exterior"]),
        ("colonia", ["colonia"]),
        ("cp", ["cp", "c.p.", "codigo postal", "código postal"]),
        ("ciudad", ["ciudad", "municipio"]),
        ("estado", ["estado"]),
        ("pais", ["pais", "país"]),
        ("direccion", ["direccion", "dirección"]),
        ("coordenadas", ["coordenadas", "coordenada", "gps", "latitud longitud"]),
        ("link_mapa", ["link mapa", "link maps", "url mapa", "google maps"]),
        ("notas", ["notas", "observaciones"]),
    ]
    cols = mapear_columnas(headers, mapeos)

    creadas = 0
    omitidas = 0
    duplicadas = 0

    for fila in range(2, ws.max_row + 1):
        cliente = valor_celda(ws, fila, cols.get("cliente", "A"))
        if not cliente:
            omitidas += 1
            continue

        entidad_id = buscar_entidad_por_id_externo(conn, cliente)
        if not entidad_id:
            entidad_id = buscar_entidad_flexible(conn, cliente)
        if not entidad_id:
            print(f"  [domicilios] No se encontró entidad para ClienteID '{cliente}'. Omitiendo.")
            omitidas += 1
            continue

        granja_id = None
        granja_nombre = valor_celda(ws, fila, cols.get("granja"))
        if granja_nombre:
            g = conn.execute(
                "SELECT id FROM crm_granjas WHERE entidad_id = ? AND lower(trim(nombre)) = ? AND activo = 1",
                (entidad_id, granja_nombre.lower()),
            ).fetchone()
            if g:
                granja_id = g["id"]

        nombre = valor_celda(ws, fila, cols.get("nombre")) or "Principal"
        calle = valor_celda(ws, fila, cols.get("calle"))
        numero = valor_celda(ws, fila, cols.get("numero"))
        colonia = valor_celda(ws, fila, cols.get("colonia"))
        cp = valor_celda(ws, fila, cols.get("cp"))
        ciudad = valor_celda(ws, fila, cols.get("ciudad"))
        estado = valor_celda(ws, fila, cols.get("estado"))
        pais = valor_celda(ws, fila, cols.get("pais"))
        direccion = valor_celda(ws, fila, cols.get("direccion"))
        coordenadas = valor_celda(ws, fila, cols.get("coordenadas"))
        link_mapa = valor_celda(ws, fila, cols.get("link_mapa"))
        notas = valor_celda(ws, fila, cols.get("notas"))

        if not direccion and any([calle, numero, colonia, ciudad, estado]):
            partes = [str(p) for p in [calle, numero, colonia, ciudad, estado, pais, cp] if p is not None]
            direccion = ", ".join(partes)

        # Solo-altas: no duplicar el mismo domicilio al re-correr el import
        duplicado = conn.execute(
            """
            SELECT 1 FROM crm_ubicaciones
            WHERE entidad_id = ? AND granja_id IS ? AND nombre = ?
              AND direccion IS NOT NULL AND direccion = ?
            LIMIT 1
            """,
            (entidad_id, granja_id, nombre, direccion),
        ).fetchone()
        if duplicado:
            duplicadas += 1
            continue

        conn.execute(
            """
            INSERT INTO crm_ubicaciones (
                entidad_id, granja_id, nombre, tipo, calle, numero, colonia, cp,
                ciudad, estado, pais, direccion, coordenadas, link_mapa, notas
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                entidad_id, granja_id, nombre, None, calle, numero, colonia, cp,
                ciudad, estado, pais, direccion, coordenadas, link_mapa, notas,
            ),
        )
        conn.commit()
        creadas += 1

    wb.close()
    return {"creadas": creadas, "duplicadas_evitadas": duplicadas, "omitidas": omitidas}


def procesar_cat_paqueterias(conn, ruta, actualizar=False):
    """Lee CAT_PAQUETERIAS. Solo-altas: evita duplicar registros al re-correr."""
    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb["CAT_PAQUETERIAS"]
    headers = obtener_columnas(ws)

    mapeos = [
        ("cliente", ["cliente", "nombre cliente", "empresa"]),
        ("paqueteria_id", ["id paqueteria", "paqueteria id", "id"]),
        ("tipo_envio", ["tipo envio", "tipo de envio", "envio"]),
        ("paqueteria", ["paqueteria", "paquetería", "empresa envio"]),
        ("ocurre_domicilio", ["ocurre domicilio", "ocurre/domicilio"]),
        ("atencion_a", ["atencion a", "atención a", "atencion"]),
        ("telefono", ["telefono", "teléfono", "tel"]),
        ("correo_guia", ["correo guia", "correo guía", "email guia"]),
        ("tipo_pago", ["tipo pago", "tipo de pago", "pago"]),
        ("facturado_a", ["facturado a", "facturado"]),
        ("status", ["status", "estatus"]),
        ("comentarios", ["comentarios", "observaciones"]),
    ]
    cols = mapear_columnas(headers, mapeos)

    creadas = 0
    omitidas = 0
    duplicadas = 0

    for fila in range(2, ws.max_row + 1):
        cliente = valor_celda(ws, fila, cols.get("cliente", "A"))
        if not cliente:
            omitidas += 1
            continue

        entidad_id = buscar_entidad_por_id_externo(conn, cliente)
        if not entidad_id:
            entidad_id = buscar_entidad_flexible(conn, cliente)
        if not entidad_id:
            print(f"  [paqueterias] No se encontró entidad para ClienteID '{cliente}'. Omitiendo.")
            omitidas += 1
            continue

        paqueteria = valor_celda(ws, fila, cols.get("paqueteria"))
        tipo_envio = valor_celda(ws, fila, cols.get("tipo_envio"))
        atencion_a = valor_celda(ws, fila, cols.get("atencion_a"))

        # Solo-altas: no duplicar la misma paquetería al re-correr el import
        duplicado = conn.execute(
            """
            SELECT 1 FROM crm_paqueterias
            WHERE entidad_id = ? AND paqueteria IS ? AND tipo_envio IS ? AND atencion_a IS ?
            LIMIT 1
            """,
            (entidad_id, paqueteria, tipo_envio, atencion_a),
        ).fetchone()
        if duplicado:
            duplicadas += 1
            continue

        conn.execute(
            """
            INSERT INTO crm_paqueterias (
                entidad_id, paqueteria_id_externo, tipo_envio, paqueteria,
                ocurre_domicilio, atencion_a, telefono, correo_guia, tipo_pago,
                facturado_a, status, comentarios
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                entidad_id,
                valor_celda(ws, fila, cols.get("paqueteria_id")),
                tipo_envio,
                paqueteria,
                valor_celda(ws, fila, cols.get("ocurre_domicilio")),
                atencion_a,
                valor_celda(ws, fila, cols.get("telefono")),
                valor_celda(ws, fila, cols.get("correo_guia")),
                valor_celda(ws, fila, cols.get("tipo_pago")),
                valor_celda(ws, fila, cols.get("facturado_a")),
                valor_celda(ws, fila, cols.get("status")) or "Activo",
                valor_celda(ws, fila, cols.get("comentarios")),
            ),
        )
        conn.commit()
        creadas += 1

    wb.close()
    return {"creadas": creadas, "duplicadas_evitadas": duplicadas, "omitidas": omitidas}


def procesar_portales(conn, ruta, actualizar=False):
    """Lee portales. Solo-altas: no actualiza portales existentes."""
    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb.active
    headers = obtener_columnas(ws)

    mapeos = [
        ("cliente", ["cliente", "nombre cliente", "empresa", "razon social", "razón social"]),
        ("portal", ["portal", "nombre portal", "pagina", "página"]),
        ("url", ["url", "link", "sitio", "direccion web"]),
        ("usuario", ["usuario", "user", "login", "usr"]),
        ("password", ["password", "contraseña", "pass", "contrasena", "contraseña portal"]),
        ("persona_apoyo", ["persona apoyo", "apoyo", "contacto apoyo"]),
        ("notas", ["notas", "observaciones"]),
    ]
    cols = mapear_columnas(headers, mapeos)

    creados = 0
    actualizados = 0
    omitidos = 0

    for fila in range(2, ws.max_row + 1):
        cliente = valor_celda(ws, fila, cols.get("cliente", "A"))
        if not cliente:
            omitidos += 1
            continue

        entidad_id = buscar_entidad_flexible(conn, cliente)
        if not entidad_id:
            print(f"  [portales] No se encontró entidad para '{cliente}'. Omitiendo.")
            omitidos += 1
            continue

        portal = valor_celda(ws, fila, cols.get("portal")) or "Portal"
        url = valor_celda(ws, fila, cols.get("url"))
        usuario = valor_celda(ws, fila, cols.get("usuario"))
        password = valor_celda(ws, fila, cols.get("password")) or ""
        persona_apoyo = valor_celda(ws, fila, cols.get("persona_apoyo"))
        notas = valor_celda(ws, fila, cols.get("notas"))

        existente = conn.execute(
            "SELECT id FROM crm_portales WHERE entidad_id = ? AND lower(trim(nombre)) = ?",
            (entidad_id, portal.lower()),
        ).fetchone()

        if existente:
            if not actualizar:
                omitidos += 1  # Ya existía: no se pisan ediciones del panel
                continue
            conn.execute(
                "UPDATE crm_portales SET url = ?, usuario = ?, password = ?, persona_apoyo = ?, notas = ? WHERE id = ?",
                (url, usuario, password, persona_apoyo, notas, existente["id"]),
            )
            actualizados += 1
        else:
            conn.execute(
                "INSERT INTO crm_portales (entidad_id, nombre, url, usuario, password, persona_apoyo, notas) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (entidad_id, portal, url, usuario, password, persona_apoyo, notas),
            )
        conn.commit()
        if not existente:
            creados += 1

    wb.close()
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos}


def procesar_contactos(conn, ruta, actualizar=False):
    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb.active
    headers = obtener_columnas(ws)

    mapeos = [
        ("cliente", ["cliente", "nombre cliente", "empresa"]),
        ("nombre", ["nombre", "contacto", "nombre contacto"]),
        ("puesto", ["puesto", "cargo"]),
        ("departamento", ["departamento", "area", "área"]),
        ("telefono", ["telefono", "teléfono", "tel"]),
        ("whatsapp", ["whatsapp", "whats", "celular"]),
        ("email", ["correo", "email", "e-mail"]),
        ("correos_facturas", ["correos facturas", "correo facturas", "facturas"]),
        ("direccion_entrega", ["direccion entrega", "dirección entrega", "entrega"]),
        ("notas", ["notas", "observaciones"]),
    ]
    cols = mapear_columnas(headers, mapeos)

    creados = 0
    omitidos = 0
    duplicados = 0

    for fila in range(2, ws.max_row + 1):
        cliente = valor_celda(ws, fila, cols.get("cliente", "A"))
        nombre = valor_celda(ws, fila, cols.get("nombre", "B"))
        if not cliente or not nombre:
            omitidos += 1
            continue

        entidad_id = buscar_entidad_id(conn, cliente)
        if not entidad_id:
            entidad_id, _ = encontrar_o_crear_entidad(conn, cliente, {"tipo": "cliente"})

        email = valor_celda(ws, fila, cols.get("email"))

        # Solo-altas: no duplicar el mismo contacto al re-correr el import
        duplicado = conn.execute(
            """
            SELECT 1 FROM crm_contactos
            WHERE entidad_id = ? AND lower(trim(nombre)) = ? AND email IS ?
            LIMIT 1
            """,
            (entidad_id, str(nombre).lower(), email),
        ).fetchone()
        if duplicado:
            duplicados += 1
            continue

        # El primer contacto de cada entidad se marca como principal si no hay otros
        tiene_principal = conn.execute(
            "SELECT 1 FROM crm_contactos WHERE entidad_id = ? AND principal = 1 LIMIT 1",
            (entidad_id,),
        ).fetchone()

        conn.execute(
            """
            INSERT INTO crm_contactos (
                entidad_id, nombre, puesto, departamento, telefono, whatsapp,
                email, correos_facturas, direccion_entrega, principal, notas
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                entidad_id,
                nombre,
                valor_celda(ws, fila, cols.get("puesto")),
                valor_celda(ws, fila, cols.get("departamento")),
                valor_celda(ws, fila, cols.get("telefono")),
                valor_celda(ws, fila, cols.get("whatsapp")),
                email,
                valor_celda(ws, fila, cols.get("correos_facturas")),
                valor_celda(ws, fila, cols.get("direccion_entrega")),
                0 if tiene_principal else 1,
                valor_celda(ws, fila, cols.get("notas")),
            ),
        )
        conn.commit()
        creados += 1

    wb.close()
    return {"creados": creados, "duplicados_evitados": duplicados, "omitidos": omitidos}


def procesar_contactos_jerarquico(conn, ruta):
    """Procesa archivos como CRM CLIENTES.xlsx donde el cliente es cabecera
    y los contactos van debajo sin repetir el nombre del cliente."""
    wb = load_workbook(ruta, data_only=True, read_only=True)
    # Usar la primera hoja que tenga filas con datos
    ws = None
    for nombre_hoja in wb.sheetnames:
        candidata = wb[nombre_hoja]
        if candidata.max_row > 1:
            ws = candidata
            break
    if ws is None:
        wb.close()
        return {"creados": 0, "omitidos": 0, "error": "Sin hojas"}
    # Encontrar fila de headers (primera fila con 'nombre' en columna A)
    fila_headers = 1
    for fila in range(1, min(6, ws.max_row + 1)):
        val_a = limpiar_valor(ws.cell(row=fila, column=1).value)
        if val_a and "nombre" in str(val_a).lower():
            fila_headers = fila
            break

    headers = {}
    for cell in ws[fila_headers]:
        if cell.value is not None:
            headers[normalizar_nombre(str(cell.value))] = cell.column_letter
    print(f"  [contactos] Hoja usada: {ws.title}, filas: {ws.max_row}, fila_headers: {fila_headers}")

    mapeos = [
        ("nombre", ["nombre", "contacto", "nombre contacto"]),
        ("departamento", ["departamento", "area", "área"]),
        ("telefono", ["numero", "número", "telefono", "teléfono", "tel"]),
        ("email", ["correo", "email", "e-mail"]),
        ("correos_facturas", ["correos para mandar facturas", "correos facturas", "correo facturas", "facturas"]),
        ("direccion_entrega", ["direccion de entrega", "dirección de entrega", "direccion entrega", "entrega"]),
        ("portal_facturas", ["portal facturas", "portal"]),
    ]
    cols = mapear_columnas(headers, mapeos)

    creados = 0
    omitidos = 0
    duplicados = 0
    cliente_actual = None
    debug_filas = 0
    cliente_actual_id = None

    def extraer_cliente(texto):
        """Extrae nombre de cliente quitando '# NUMERO' al final."""
        if not texto:
            return None
        texto = str(texto).strip()
        # Quitar parte como '# VARIOS' o '# 60148'
        texto = re.sub(r"\s*#\s*\d+.*$", "", texto).strip()
        return texto if texto else None

    for fila in range(fila_headers + 1, ws.max_row + 1):
        valores = [ws.cell(row=fila, column=col).value for col in range(1, ws.max_column + 1)]
        primera_celda = limpiar_valor(valores[0])
        resto = [limpiar_valor(v) for v in valores[1:]]

        # Si solo la primera columna tiene valor, es un cliente cabecera
        if primera_celda and not any(resto):
            cliente_actual = extraer_cliente(primera_celda)
            cliente_actual_id = buscar_entidad_flexible(conn, cliente_actual) if cliente_actual else None
            if cliente_actual_id is None and cliente_actual:
                print(f"  [contactos] No se encontró entidad para '{cliente_actual}'. Omitiendo contactos.")
            continue

        if not primera_celda:
            continue

        if cliente_actual_id is None:
            omitidos += 1
            continue

        nombre = primera_celda
        email = valor_celda(ws, fila, cols.get("email"))

        # Solo-altas: no duplicar el mismo contacto al re-correr el import
        duplicado = conn.execute(
            """
            SELECT 1 FROM crm_contactos
            WHERE entidad_id = ? AND lower(trim(nombre)) = ? AND email IS ?
            LIMIT 1
            """,
            (cliente_actual_id, str(nombre).lower(), email),
        ).fetchone()
        if duplicado:
            duplicados += 1
            continue

        tiene_principal = conn.execute(
            "SELECT 1 FROM crm_contactos WHERE entidad_id = ? AND principal = 1 LIMIT 1",
            (cliente_actual_id,),
        ).fetchone()

        conn.execute(
            """
            INSERT INTO crm_contactos (
                entidad_id, nombre, puesto, departamento, telefono, whatsapp,
                email, correos_facturas, direccion_entrega, principal, notas
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                cliente_actual_id,
                nombre,
                None,
                valor_celda(ws, fila, cols.get("departamento")),
                valor_celda(ws, fila, cols.get("telefono")),
                None,
                email,
                valor_celda(ws, fila, cols.get("correos_facturas")),
                valor_celda(ws, fila, cols.get("direccion_entrega")),
                0 if tiene_principal else 1,
                valor_celda(ws, fila, cols.get("portal_facturas")),
            ),
        )
        conn.commit()
        creados += 1

    wb.close()
    return {"creados": creados, "duplicados_evitados": duplicados, "omitidos": omitidos}


def procesar_descuentos(conn, ruta, actualizar=False):
    """Procesa CRM DESCUENTOS CLIENTES.xlsx con dos tablas lado a lado:
    A-B: CLIENTE, DIAS DE CREDITO
    E-K: CLIENTE, LUBING, ROXELL, CHORE T., SBM, FANCOM, GROWER

    Solo-altas (default): no pisa días de crédito ni descuentos ya
    existentes (los edita el panel admin). Con actualizar=True recupera
    el comportamiento anterior.
    """
    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb.active

    creados = 0
    actualizados = 0
    omitidos = 0

    MARCAS = ["LUBING", "ROXELL", "CHORE TIME", "SBM", "FANCOM", "GROWER"]

    for fila in range(5, ws.max_row + 1):
        # Tabla 1: días de crédito (columnas A-B)
        cliente_credito = limpiar_valor(ws.cell(row=fila, column=1).value)
        dias_credito = limpiar_valor(ws.cell(row=fila, column=2).value)

        if cliente_credito and dias_credito:
            entidad_id = buscar_entidad_flexible(conn, cliente_credito)
            if not entidad_id:
                print(f"  [descuentos] No se encontró entidad para '{cliente_credito}'. Omitiendo crédito.")
                omitidos += 1
            elif not actualizar:
                omitidos += 1  # Ya existía: no se pisan ediciones del panel
            else:
                conn.execute(
                    "UPDATE crm_entidades SET dias_credito = ?, updated_at = ? WHERE id = ?",
                    (str(dias_credito), datetime.utcnow().isoformat(), entidad_id),
                )
                conn.commit()
                actualizados += 1

        # Tabla 2: descuentos por marca (columnas E-K)
        cliente_desc = limpiar_valor(ws.cell(row=fila, column=5).value)
        if not cliente_desc:
            continue

        entidad_id = buscar_entidad_flexible(conn, cliente_desc)
        if not entidad_id:
            print(f"  [descuentos] No se encontró entidad para '{cliente_desc}'. Omitiendo descuentos.")
            omitidos += 1
            continue

        for i, marca in enumerate(MARCAS):
            col = 6 + i  # F=6, G=7, ...
            descuento_val = limpiar_valor(ws.cell(row=fila, column=col).value)
            if descuento_val:
                existente = conn.execute(
                    "SELECT id FROM crm_descuentos WHERE entidad_id = ? AND upper(marca) = ?",
                    (entidad_id, marca.upper()),
                ).fetchone()
                if existente:
                    if not actualizar:
                        omitidos += 1  # Ya existía: no se pisan ediciones del panel
                        continue
                    conn.execute(
                        "UPDATE crm_descuentos SET descuento = ? WHERE id = ?",
                        (str(descuento_val), existente["id"]),
                    )
                    actualizados += 1
                else:
                    conn.execute(
                        "INSERT INTO crm_descuentos (entidad_id, marca, descuento, notas) VALUES (?, ?, ?, ?)",
                        (entidad_id, marca, str(descuento_val), None),
                    )
                    creados += 1
                conn.commit()

    wb.close()
    return {"creados_descuentos": creados, "actualizados": actualizados, "omitidos": omitidos}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def encontrar_archivo(directorios, nombres):
    """Busca un archivo en varios directorios con varios nombres posibles."""
    for directorio in directorios:
        for nombre in nombres:
            ruta = directorio / nombre
            if ruta.exists():
                return ruta
    return None


def main():
    parser = argparse.ArgumentParser(description="Importa datos de CRM desde Excel a SQLite")
    parser.add_argument(
        "--dir",
        type=Path,
        default=None,
        help="Directorio donde se encuentran los archivos Excel (por defecto Y:/CRM´S)",
    )
    parser.add_argument(
        "--actualizar",
        action="store_true",
        help=(
            "Pisa los datos existentes con lo que traiga el Excel (comportamiento "
            "anterior). Por defecto el import es SOLO-ALTAS: respeta las ediciones "
            "hechas en el panel admin, que es la fuente de verdad del CRM."
        ),
    )
    args = parser.parse_args()

    actualizar = args.actualizar

    base_dir = args.dir
    if not base_dir:
        base_dir = DEFAULT_DIR if DEFAULT_DIR.exists() else ALT_DIR

    if not base_dir.exists():
        print(f"ERROR: No se encontró el directorio {base_dir}")
        print("Usa --dir para especificar la ruta correcta.")
        sys.exit(1)

    # Directorios donde buscar archivos adicionales
    dirs = [base_dir, Path("Y:/1 - CONTROL DE ALMACEN")]

    print(f"Importando desde: {base_dir}")
    print(f"Modo: {'ACTUALIZAR (pisa datos existentes)' if actualizar else 'SOLO-ALTAS (respeta ediciones del panel admin)'}")
    print("=" * 60)

    archivos = {
        "final_3p": encontrar_archivo(dirs, ["CRM´S FINAL 3P.xlsm", "CRM'S FINAL 3P.xlsm"]),
        "portales": encontrar_archivo(dirs, ["CRM PORTALES CLIENTES SUBIR FACTURAS.xlsx"]),
        "clientes": encontrar_archivo(dirs, ["CRM CLIENTES.xlsx"]),
        "clientes_joan": encontrar_archivo(dirs, ["CRM Clientes Joan.xlsx"]),
        "descuentos": encontrar_archivo(dirs, ["CRM DESCUENTOS CLIENTES.xlsx"]),
    }

    resultados = {}

    with users_connection() as conn:
        if archivos["final_3p"] and archivos["final_3p"].exists():
            print(f"\n[1/5] Procesando {archivos['final_3p'].name}")
            resultados["cat_clientes"] = procesar_cat_clientes(conn, archivos["final_3p"], actualizar)
            resultados["cat_granjas"] = procesar_cat_granjas(conn, archivos["final_3p"], actualizar)
            resultados["cat_domicilios"] = procesar_cat_domicilios(conn, archivos["final_3p"], actualizar)
            resultados["cat_paqueterias"] = procesar_cat_paqueterias(conn, archivos["final_3p"], actualizar)
        else:
            print(f"\n[1/5] No se encontró CRM´S FINAL 3P.xlsm, se omite.")

        if archivos["portales"] and archivos["portales"].exists():
            print(f"\n[2/5] Procesando {archivos['portales'].name}")
            resultados["portales"] = procesar_portales(conn, archivos["portales"], actualizar)
        else:
            print(f"\n[2/5] No se encontró CRM PORTALES CLIENTES SUBIR FACTURAS.xlsx, se omite.")

        if archivos["clientes"] and archivos["clientes"].exists():
            print(f"\n[3/5] Procesando {archivos['clientes'].name}")
            resultados["contactos"] = procesar_contactos_jerarquico(conn, archivos["clientes"], actualizar)
        else:
            print(f"\n[3/5] No se encontró archivo de contactos, se omite.")

        if archivos["clientes_joan"] and archivos["clientes_joan"].exists():
            print(f"\n[3b/5] Procesando {archivos['clientes_joan'].name}")
            resultados["contactos_joan"] = procesar_contactos(conn, archivos["clientes_joan"], actualizar)
        else:
            print(f"\n[3b/5] No se encontró CRM Clientes Joan.xlsx, se omite.")

        if archivos["descuentos"] and archivos["descuentos"].exists():
            print(f"\n[4/5] Procesando {archivos['descuentos'].name}")
            resultados["descuentos"] = procesar_descuentos(conn, archivos["descuentos"], actualizar)
        else:
            print(f"\n[4/5] No se encontró CRM DESCUENTOS CLIENTES.xlsx, se omite.")

        print("\n[5/5] Recalculando conteos...")
        counts = dict(
            conn.execute(
                """
                SELECT
                    (SELECT COUNT(*) FROM crm_entidades WHERE activo = 1) AS entidades,
                    (SELECT COUNT(*) FROM crm_contactos c
                     JOIN crm_entidades e ON e.id = c.entidad_id WHERE e.activo = 1) AS contactos,
                    (SELECT COUNT(*) FROM crm_granjas g
                     JOIN crm_entidades e ON e.id = g.entidad_id WHERE e.activo = 1 AND g.activo = 1) AS granjas,
                    (SELECT COUNT(*) FROM crm_ubicaciones u
                     JOIN crm_entidades e ON e.id = u.entidad_id WHERE e.activo = 1) AS ubicaciones,
                    (SELECT COUNT(*) FROM crm_paqueterias p
                     JOIN crm_entidades e ON e.id = p.entidad_id WHERE e.activo = 1) AS paqueterias,
                    (SELECT COUNT(*) FROM crm_portales p
                     JOIN crm_entidades e ON e.id = p.entidad_id WHERE e.activo = 1) AS portales,
                    (SELECT COUNT(*) FROM crm_descuentos d
                     JOIN crm_entidades e ON e.id = d.entidad_id WHERE e.activo = 1) AS descuentos
                """
            ).fetchone()
        )

    print("\n" + "=" * 60)
    print("RESUMEN DE IMPORTACIÓN")
    print("=" * 60)
    for seccion, res in resultados.items():
        print(f"\n{seccion}:")
        for k, v in res.items():
            print(f"  - {k}: {v}")

    print("\nTotales en base de datos:")
    for k, v in counts.items():
        print(f"  - {k}: {v}")
    print("\nImportación finalizada.")


if __name__ == "__main__":
    main()
