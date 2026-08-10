from pathlib import Path
from openpyxl import load_workbook

excel_path = Path("Y:/ALMACEN/Mejora Continua ALMACEN/Nuevo Control de Almacen/BASE DE DATOS/BD_ALMACEN_3P.xlsx")

if not excel_path.exists():
    print(f"No existe: {excel_path}")
    exit(1)

print(f"Archivo: {excel_path}")
print(f"Tamaño: {excel_path.stat().st_size} bytes")
print()

try:
    wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    print("Hojas:", wb.sheetnames)
    print()

    for sheet_name in wb.sheetnames:
        print(f"=== Hoja: {sheet_name} ===")
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        print(f"Columnas ({len(headers)}): {headers}")

        rows = ws.iter_rows(min_row=2, values_only=True)
        for i, row in enumerate(rows):
            if i >= 3:
                break
            if not row or all(v is None for v in row):
                continue
            print(f"Fila {i+2}: {dict(zip(headers, row))}")
        print()

    wb.close()
except Exception as e:
    print(f"ERROR: {e}")
