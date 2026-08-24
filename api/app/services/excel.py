"""Servicios de lectura de archivos Excel compartidos entre routers."""

import datetime
import json
import time
from pathlib import Path

from openpyxl import load_workbook

from app.config import get_settings


# ---------------------------------------------------------------------------
# Utilidades comunes
# ---------------------------------------------------------------------------


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def read_excel_sheet(wb, sheet_name):
    """Lee una hoja de openpyxl como lista de diccionados (header -> valor)."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [normalize_text(cell.value) for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        item = {}
        for h, v in zip(headers, row):
            if h:
                item[h] = v
        data.append(item)
    return data


def to_date(value):
    """Normaliza un valor de fecha de Excel a datetime.date."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


# ---------------------------------------------------------------------------
# Vales
# ---------------------------------------------------------------------------


def _vales_excel_path() -> Path:
    return Path(get_settings().vales_excel_path)


def _load_vales_excel():
    excel_path = _vales_excel_path()
    if not excel_path.exists():
        return [], []
    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception:
        return [], []
    try:
        cabeceras = read_excel_sheet(wb, "VALES")
        detalles = read_excel_sheet(wb, "DETALLE_VALES")
    finally:
        wb.close()
    return cabeceras, detalles


def _index_cabeceras_by_folio(cabeceras):
    by_folio = {}
    for c in cabeceras:
        folio = c.get("FOLIO_VALE")
        if folio is not None:
            by_folio[str(folio).strip()] = c
    return by_folio


def _cantidad_viva(detail_row, cabecera):
    status_val = normalize_text(cabecera.get("STATUS") or detail_row.get("STATUS")).upper()
    cantidad_viva = detail_row.get("CANTIDAD_VIVA", 0) or 0
    try:
        cantidad_viva = float(cantidad_viva)
    except (ValueError, TypeError):
        cantidad_viva = 0
    if status_val == "CERRADO" or cantidad_viva <= 0:
        return 0, status_val
    return cantidad_viva, status_val


def get_material_en_vales_by_code():
    """Devuelve dict {codigo: cantidad_viva_total} desde el Excel de vales."""
    cabeceras, detalles = _load_vales_excel()
    cabeceras_by_folio = _index_cabeceras_by_folio(cabeceras)
    material = {}
    for d in detalles:
        folio = normalize_text(d.get("FOLIO_VALE"))
        cab = cabeceras_by_folio.get(folio, {})
        cantidad_viva, _ = _cantidad_viva(d, cab)
        if cantidad_viva <= 0:
            continue
        codigo = normalize_text(d.get("CODIGO"))
        if codigo:
            material[codigo] = material.get(codigo, 0) + cantidad_viva
    return material


def get_vales_abiertos_count():
    """Cuenta folios de vale distintos con cantidad viva > 0."""
    cabeceras, detalles = _load_vales_excel()
    cabeceras_by_folio = _index_cabeceras_by_folio(cabeceras)
    folios_abiertos = set()
    for d in detalles:
        folio = normalize_text(d.get("FOLIO_VALE"))
        cab = cabeceras_by_folio.get(folio, {})
        cantidad_viva, _ = _cantidad_viva(d, cab)
        if cantidad_viva <= 0:
            continue
        folios_abiertos.add(folio)
    return len(folios_abiertos)


# ---------------------------------------------------------------------------
# Pedidos vivos
# ---------------------------------------------------------------------------


def get_pedidos_vivos_excel(busqueda: str = "", limit: int | None = None):
    """Lee los pedidos vivos reales desde el Excel de pendientes por facturar."""
    settings = get_settings()
    excel_path = Path(settings.pedidos_pendientes_facturar_excel_path)

    if not excel_path.exists():
        return []

    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception:
        return []

    try:
        cabeceras = read_excel_sheet(wb, "PEDIDOS")
        detalles = read_excel_sheet(wb, "DETALLE_PEDIDOS")
    finally:
        wb.close()

    agg = {}
    for d in detalles:
        folio = normalize_text(d.get("FOLIO_PEDIDO"))
        if not folio:
            continue
        if folio not in agg:
            agg[folio] = {
                "importe_total": 0.0,
                "total_facturado": 0.0,
                "saldo_pendiente": 0.0,
            }
        cant_pedida = d.get("CANT_PEDIDA", 0) or 0
        precio_unitario = d.get("PRECIO_UNITARIO", 0) or 0
        total_facturado = d.get("TOTAL_FACTURADO", 0) or 0
        pendiente_facturar = d.get("PENDIENTE_FACTURAR", 0) or 0
        try:
            agg[folio]["importe_total"] += float(cant_pedida) * float(precio_unitario)
            agg[folio]["total_facturado"] += float(total_facturado)
            agg[folio]["saldo_pendiente"] += float(pendiente_facturar)
        except (ValueError, TypeError):
            continue

    busqueda_lower = busqueda.lower().strip()
    resultados = []
    for c in cabeceras:
        folio = normalize_text(c.get("FOLIO_PEDIDO"))
        if not folio:
            continue
        cliente = normalize_text(c.get("CLIENTE"))
        if busqueda_lower and (
            busqueda_lower not in folio.lower()
            and busqueda_lower not in cliente.lower()
        ):
            continue

        fecha = c.get("FECHA_PEDIDO")
        dias_pendiente = None
        if isinstance(fecha, datetime.datetime):
            dias_pendiente = (datetime.date.today() - fecha.date()).days
        elif isinstance(fecha, datetime.date):
            dias_pendiente = (datetime.date.today() - fecha).days

        totales = agg.get(folio, {})
        saldo_pendiente = totales.get("saldo_pendiente", 0.0)
        importe_total = totales.get("importe_total", 0.0)
        total_facturado = totales.get("total_facturado", 0.0)

        if saldo_pendiente <= 0.01:
            continue

        if total_facturado > 0.01:
            estado_calculado = "Parcial"
        else:
            estado_calculado = "Pendiente"

        moneda_val = normalize_text(c.get("MONEDA") or c.get("MONEDA_PEDIDO") or "MXN").upper()
        if moneda_val not in ("USD", "MXN", "PESOS", "DOLARES", "DÓLARES"):
            moneda_val = "MXN"
        if moneda_val in ("PESOS",):
            moneda_val = "MXN"
        if moneda_val in ("DOLARES", "DÓLARES"):
            moneda_val = "USD"

        resultados.append({
            "folio": folio,
            "cliente": cliente,
            "fecha": fecha.isoformat() if hasattr(fecha, "isoformat") else fecha,
            "importe_total": importe_total,
            "total_facturado": total_facturado,
            "saldo_pendiente": saldo_pendiente,
            "estado": estado_calculado,
            "estado_original": normalize_text(c.get("STATUS_GENERAL")),
            "moneda": moneda_val,
            "dias_pendiente": dias_pendiente,
        })

    resultados.sort(key=lambda x: x["fecha"] or "", reverse=True)
    if limit:
        resultados = resultados[:limit]
    return resultados


def get_pedido_detalle_excel(folio_pedido: str = ""):
    """Lee el detalle de un pedido desde el Excel de pendientes por facturar.

    Funciona como fallback cuando la vista v_seguimiento_documental no está
    disponible en PostgreSQL.
    """
    settings = get_settings()
    excel_path = Path(settings.pedidos_pendientes_facturar_excel_path)

    if not excel_path.exists():
        return []

    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception:
        return []

    try:
        detalles = read_excel_sheet(wb, "DETALLE_PEDIDOS")
    finally:
        wb.close()

    folio_lower = folio_pedido.lower().strip()
    resultados = []
    for d in detalles:
        folio = normalize_text(d.get("FOLIO_PEDIDO"))
        if folio_lower and folio_lower != folio.lower():
            continue

        cant_pedida = _to_float(d.get("CANT_PEDIDA"))
        precio_unitario = _to_float(d.get("PRECIO_UNITARIO"))
        total_facturado = _to_float(d.get("TOTAL_FACTURADO"))
        pendiente_facturar = _to_float(d.get("PENDIENTE_FACTURAR"))

        resultados.append({
            "folio_pedido": folio,
            "fecha_pedido": None,
            "cliente": normalize_text(d.get("CLIENTE")),
            "codigo": normalize_text(d.get("CODIGO") or d.get("CODIGO_PRODUCTO")),
            "descripcion": normalize_text(d.get("DESCRIPCION") or d.get("DESCRIPCION_PRODUCTO")),
            "cantidad_pedido": cant_pedida,
            "folio_remision": normalize_text(d.get("FOLIO_REMISION")),
            "cantidad_remision": _to_float(d.get("CANT_REMISION")),
            "folio_factura": normalize_text(d.get("FOLIO_FACTURA")),
            "cantidad_factura": _to_float(d.get("CANT_FACTURA")),
            "estatus_linea": normalize_text(d.get("ESTATUS_LINEA") or d.get("STATUS")),
            "precio_unitario": precio_unitario,
            "total_facturado": total_facturado,
            "pendiente_facturar": pendiente_facturar,
        })

    return resultados


# ---------------------------------------------------------------------------
# Fotos de productos
# ---------------------------------------------------------------------------

_FOTOS_CACHE = {"map": None, "ts": 0}
_FOTOS_TTL_SECONDS = 300


def get_fotos_map():
    """Lee la hoja FOTOS_PRODUCTOS del Excel de almacén y devuelve un dict.

    Si el archivo no existe o no se puede abrir, devuelve dict vacío.
    La última versión del mapa se conserva en memoria por _FOTOS_TTL_SECONDS.
    """
    global _FOTOS_CACHE
    now = time.time()
    if _FOTOS_CACHE["map"] is not None and (now - _FOTOS_CACHE["ts"]) < _FOTOS_TTL_SECONDS:
        return _FOTOS_CACHE["map"]

    excel_path = _vales_excel_path()
    if not excel_path.exists():
        _FOTOS_CACHE = {"map": {}, "ts": now}
        return {}

    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception:
        _FOTOS_CACHE = {"map": {}, "ts": now}
        return {}

    try:
        fotos = {}
        if "FOTOS_PRODUCTOS" not in wb.sheetnames:
            _FOTOS_CACHE = {"map": {}, "ts": now}
            return {}
        ws = wb["FOTOS_PRODUCTOS"]
        headers = [normalize_text(cell.value) for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue
            item = {}
            for h, v in zip(headers, row):
                if h:
                    item[h] = v
            codigo = normalize_text(item.get("CODIGO"))
            ruta = normalize_text(item.get("RUTA_FOTO"))
            if codigo and ruta:
                fotos[codigo] = ruta
    finally:
        wb.close()

    _FOTOS_CACHE = {"map": fotos, "ts": now}
    return fotos


# Alias interno para compatibilidad con imports anteriores
_get_fotos_map = get_fotos_map


# ---------------------------------------------------------------------------
# Historial de ventas / facturación
# ---------------------------------------------------------------------------

_CACHE_TTL_SECONDS = 30 * 60  # 30 minutos
_historial_cache = {
    "timestamp": 0,
    "excel_mtime": 0,
    "filas": [],
    "clientes": [],
    "codigos": [],
}


def _to_float(value):
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _iso_date(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value) if value else None


def _read_seguimiento_documental(wb):
    """Lee la hoja 'Seguimiento_Documental' cuyos headers están en la fila 4."""
    sheet_name = "Seguimiento_Documental"
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)

    for _ in range(3):
        next(rows_iter, None)

    headers = [normalize_text(cell) for cell in next(rows_iter, [])]
    if not headers:
        return []

    data = []
    for row in rows_iter:
        if not row or all(v is None for v in row):
            continue
        item = {}
        for h, v in zip(headers, row):
            if h:
                item[h] = v
        data.append(item)
    return data


def _build_historial_item(item):
    """Convierte una fila cruda del Excel en el diccionario que expone la API."""
    cantidad = _to_float(item.get("Cantidad"))
    precio_unitario = _to_float(item.get("Precio Unitario"))
    importe_partida = _to_float(item.get("Importe Partida"))
    tipo_cambio = _to_float(item.get("Tipo de Cambio"))

    return {
        "cliente": normalize_text(item.get("Cliente Pedido")),
        "codigo": normalize_text(item.get("Codigo Producto")),
        "descripcion": normalize_text(item.get("Descripcion Producto")),
        "cantidad": cantidad,
        "precio_unitario": precio_unitario,
        "importe_partida": importe_partida,
        "moneda": normalize_text(item.get("Moneda")).upper() or "MXN",
        "tipo_cambio": tipo_cambio,
        "almacen": normalize_text(item.get("Nombre Almacen Linea")),
        "folio_factura": normalize_text(item.get("Folio Factura")),
        "folio_pedido": normalize_text(item.get("Folio Pedido")),
        "fecha_factura": _iso_date(item.get("Fecha Factura")),
        "fecha_pedido": _iso_date(item.get("Fecha Pedido")),
    }


def _cache_file_path(excel_path: Path) -> Path:
    cache_dir = Path(__file__).resolve().parent.parent / "data" / "cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / f"{excel_path.stem}_historial_cache.json"


def _save_cache_to_disk(excel_path: Path):
    try:
        cache_path = _cache_file_path(excel_path)
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(_historial_cache, f, ensure_ascii=False, default=str)
    except Exception:
        pass


def _load_cache_from_disk(excel_path: Path) -> bool:
    global _historial_cache
    cache_path = _cache_file_path(excel_path)
    if not cache_path.exists():
        return False

    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            cached = json.load(f)

        excel_mtime = excel_path.stat().st_mtime
        if cached.get("excel_mtime") != excel_mtime:
            return False

        _historial_cache = cached
        _historial_cache["timestamp"] = time.time()
        return True
    except Exception:
        return False


def _load_historial_cache():
    """Lee el Excel de facturación y guarda en caché las filas de factura y metadatos."""
    global _historial_cache
    settings = get_settings()
    excel_path = Path(settings.ventas_facturacion_excel_path)

    if not excel_path.exists():
        _historial_cache = {
            "timestamp": time.time(),
            "excel_mtime": 0,
            "filas": [],
            "clientes": [],
            "codigos": [],
        }
        return

    if _load_cache_from_disk(excel_path):
        return

    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception:
        return

    try:
        filas_crudas = _read_seguimiento_documental(wb)
    finally:
        wb.close()

    filas = []
    clientes_set = set()
    codigos_set = set()

    for item in filas_crudas:
        tipo_fila = normalize_text(item.get("Tipo de Fila")).lower()
        if "factura" not in tipo_fila:
            continue
        fila = _build_historial_item(item)
        filas.append(fila)
        if fila["cliente"]:
            clientes_set.add(fila["cliente"])
        if fila["codigo"]:
            codigos_set.add(fila["codigo"])

    filas.sort(key=lambda x: (x["cliente"] or "", x["fecha_factura"] or "", x["codigo"] or ""))

    _historial_cache = {
        "timestamp": time.time(),
        "excel_mtime": excel_path.stat().st_mtime,
        "filas": filas,
        "clientes": sorted(clientes_set),
        "codigos": sorted(codigos_set),
    }

    _save_cache_to_disk(excel_path)


def precargar_historial_cache():
    """Fuerza la carga inicial del caché del historial de ventas."""
    _load_historial_cache()


def _get_cached_historial():
    """Devuelve la caché del historial, recargándola si expiró o cambió el Excel."""
    settings = get_settings()
    excel_path = Path(settings.ventas_facturacion_excel_path)

    needs_reload = False
    if not _historial_cache["filas"]:
        needs_reload = True
    elif time.time() - _historial_cache["timestamp"] > _CACHE_TTL_SECONDS:
        needs_reload = True
    elif excel_path.exists() and excel_path.stat().st_mtime != _historial_cache["excel_mtime"]:
        needs_reload = True

    if needs_reload:
        _load_historial_cache()

    return _historial_cache


def get_historial_filas():
    return _get_cached_historial()["filas"]


def get_historial_clientes():
    return _get_cached_historial()["clientes"]


def get_historial_codigos():
    return _get_cached_historial()["codigos"]
