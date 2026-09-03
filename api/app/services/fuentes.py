"""Estado y frescura de las fuentes de datos externas (Excel y espejo SAE).

Solo lectura: este módulo NUNCA escribe en los archivos Excel origen.
Alimenta GET /health/datos y la alerta de fuentes en app/analytics/alertas.py.
"""

import logging
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from app.config import get_settings
from app.services.excel import normalize_text

logger = logging.getLogger(__name__)

# Edad máxima aceptable del mtime por fuente (horarios de operación reales).
# None = no se evalúa antigüedad (fuente de cambio raro).
MAX_AGE_HORAS = {
    "vales": 24,
    "pedidos": 48,
    "ventas_facturacion": 72,
    "cotizador_vendedores": None,
    "san_antonio": 24 * 8,  # semanal
}

# Hojas y columnas que el código espera de cada fuente (mitiga R3:
# renombres de hojas/columnas rompen en silencio). Las columnas críticas
# son las que los routers leen directamente con .get().
FUENTES_EXCEL = {
    "vales": {
        "setting": "vales_excel_path",
        "sheets": {
            "VALES": ["FOLIO_VALE", "STATUS"],
            "DETALLE_VALES": ["FOLIO_VALE", "CODIGO", "CANTIDAD_VIVA"],
            "FOTOS_PRODUCTOS": ["CODIGO", "RUTA_FOTO"],
        },
    },
    "pedidos": {
        "setting": "pedidos_pendientes_facturar_excel_path",
        "sheets": {
            "PEDIDOS": ["FOLIO_PEDIDO", "CLIENTE", "FECHA_PEDIDO"],
            "DETALLE_PEDIDOS": [
                "FOLIO_PEDIDO",
                "CODIGO",
                "CANT_PEDIDA",
                "PRECIO_UNITARIO",
                "TOTAL_FACTURADO",
                "PENDIENTE_FACTURAR",
            ],
        },
    },
    "ventas_facturacion": {
        "setting": "ventas_facturacion_excel_path",
        # La hoja Seguimiento_Documental tiene sus headers en la fila 4.
        "header_row": 4,
        "sheets": {
            "Seguimiento_Documental": ["Tipo de Fila", "Folio Factura", "Codigo Producto"],
        },
    },
    "cotizador_vendedores": {
        "setting": "cotizador_vendedores_excel_path",
        "sheets": {
            "FIRMAS": ["NOMBRE", "FIRMA"],
        },
    },
    "san_antonio": {
        "setting": "san_antonio_excel_path",
        "sheets": {
            "OC_CABECERA": ["Folio", "NoPedido", "EstadoOC"],
            "OC_PARTIDAS": ["Folio", "Codigo", "CantidadPedido"],
        },
    },
}


def _headers_de_hoja(ws, header_row: int) -> list[str]:
    fila = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    if not fila:
        return []
    return [normalize_text(v) for v in fila[0]]


def _revisar_excel(nombre: str, spec: dict) -> dict:
    ruta = Path(getattr(get_settings(), spec["setting"]))
    base = {
        "fuente": nombre,
        "ruta": str(ruta),
        "existe": False,
        "mtime": None,
        "edad_horas": None,
        "tamano_bytes": None,
        "filas": 0,
        "estado": "inaccesible",
        "detalle": "",
    }
    if not ruta.exists():
        base["detalle"] = "El archivo no existe o la unidad de red no está disponible"
        return base

    stat = ruta.stat()
    base["existe"] = True
    base["tamano_bytes"] = stat.st_size
    base["mtime"] = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
    base["edad_horas"] = round(
        (datetime.now(timezone.utc) - datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)).total_seconds() / 3600,
        1,
    )

    try:
        wb = load_workbook(filename=str(ruta), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        base["estado"] = "inaccesible"
        base["detalle"] = f"No se pudo abrir: {exc}"
        return base

    try:
        header_row = spec.get("header_row", 1)
        problemas = []
        total_filas = 0
        for sheet, columnas_esperadas in spec["sheets"].items():
            if sheet not in wb.sheetnames:
                problemas.append(f"hoja '{sheet}' no existe")
                continue
            ws = wb[sheet]
            headers = set(_headers_de_hoja(ws, header_row))
            faltantes = [c for c in columnas_esperadas if c not in headers]
            if faltantes:
                problemas.append(f"hoja '{sheet}': columnas faltantes {faltantes}")
            # Restar las filas de encabezado para aproximar datos reales.
            total_filas += max((ws.max_row or 1) - header_row, 0)

        base["filas"] = total_filas
        if problemas:
            base["estado"] = "esquema_invalido"
            base["detalle"] = "; ".join(problemas)
        elif total_filas <= 0:
            base["estado"] = "vacio"
            base["detalle"] = "Sin filas de datos"
        else:
            base["estado"] = "ok"
            base["detalle"] = "OK"
        return base
    except Exception as exc:  # noqa: BLE001
        base["estado"] = "inaccesible"
        base["detalle"] = f"Error al leer: {exc}"
        return base
    finally:
        wb.close()


def _revisar_sae() -> dict:
    from app.database import postgres_cursor

    tablas = ["sae_existencias", "sae_movimientos_inventario"]
    resultado = {"fuente": "sae_postgres", "tablas": {}, "estado": "ok", "detalle": "OK"}
    try:
        with postgres_cursor() as cur:
            for tabla in tablas:
                cur.execute(f"SELECT MAX(updated_at) AS max_upd, COUNT(*) AS n FROM {tabla}")
                fila = dict(cur.fetchone())
                resultado["tablas"][tabla] = fila
                if not fila.get("max_upd"):
                    resultado["estado"] = "vacio"
    except Exception as exc:  # noqa: BLE001
        resultado["estado"] = "inaccesible"
        resultado["detalle"] = f"No se pudo consultar PostgreSQL: {exc}"
    return resultado


def estado_fuentes() -> dict:
    """Devuelve el estado completo de todas las fuentes de datos externas."""
    excel = {nombre: _revisar_excel(nombre, spec) for nombre, spec in FUENTES_EXCEL.items()}
    return {
        "revisado_en": datetime.now(timezone.utc).isoformat(),
        "excel": excel,
        "sae": _revisar_sae(),
    }
