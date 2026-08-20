import datetime
import json
import logging
import sqlite3
from io import BytesIO
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.auth.dependencies import get_current_user
from app.config import get_settings
from app.ventas.router import _get_cached_historial

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/cotizaciones", tags=["cotizaciones"])


def _cotizaciones_db_path() -> Path:
    settings = get_settings()
    base = Path(settings.users_db_path).resolve().parent
    base.mkdir(parents=True, exist_ok=True)
    return base / "cotizaciones.db"


def _column_exists(conn, table, column):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    return any(row[1] == column for row in cur.fetchall())


def _init_cotizaciones_db():
    path = _cotizaciones_db_path()
    conn = sqlite3.connect(str(path))
    try:
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS cotizaciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                folio TEXT NOT NULL UNIQUE,
                cliente TEXT NOT NULL,
                atencion TEXT,
                moneda TEXT NOT NULL DEFAULT 'USD',
                condiciones TEXT,
                tiempo_entrega TEXT,
                leyenda_envio TEXT,
                con_descuento INTEGER DEFAULT 0,
                con_stock_leon INTEGER DEFAULT 0,
                usuario_email TEXT,
                usuario_nombre TEXT,
                vendedor TEXT,
                subtotal REAL DEFAULT 0,
                iva REAL DEFAULT 0,
                total REAL DEFAULT 0,
                fecha TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS cotizacion_lineas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cotizacion_id INTEGER NOT NULL,
                codigo TEXT,
                descripcion TEXT,
                almacen TEXT,
                cantidad REAL DEFAULT 1,
                precio_unitario REAL DEFAULT 0,
                descuento_pct REAL DEFAULT 0,
                precio_con_descuento REAL DEFAULT 0,
                total_linea REAL DEFAULT 0,
                stock_leon INTEGER DEFAULT 0,
                FOREIGN KEY (cotizacion_id) REFERENCES cotizaciones(id) ON DELETE CASCADE
            )
            """
        )
        # Migraciones para tablas existentes
        if not _column_exists(conn, "cotizaciones", "leyenda_envio"):
            cur.execute("ALTER TABLE cotizaciones ADD COLUMN leyenda_envio TEXT")
        if not _column_exists(conn, "cotizaciones", "con_descuento"):
            cur.execute("ALTER TABLE cotizaciones ADD COLUMN con_descuento INTEGER DEFAULT 0")
        if not _column_exists(conn, "cotizaciones", "con_stock_leon"):
            cur.execute("ALTER TABLE cotizaciones ADD COLUMN con_stock_leon INTEGER DEFAULT 0")
        if not _column_exists(conn, "cotizaciones", "vendedor"):
            cur.execute("ALTER TABLE cotizaciones ADD COLUMN vendedor TEXT")
        if not _column_exists(conn, "cotizacion_lineas", "stock_leon"):
            cur.execute("ALTER TABLE cotizacion_lineas ADD COLUMN stock_leon INTEGER DEFAULT 0")
        # Eliminar con_envio si existe (ya no se usa)
        conn.commit()
    finally:
        conn.close()


def _db_connection():
    _init_cotizaciones_db()
    conn = sqlite3.connect(str(_cotizaciones_db_path()))
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()


@router.get("/precio-referencia")
def precio_referencia(
    codigo: str = Query(...),
    cliente: str = Query(""),
    user: dict = Depends(get_current_user),
):
    """Devuelve el último precio unitario facturado para un código.

    Prioriza el precio del cliente dado; si no existe, devuelve el último precio general.
    También devuelve la descripción y almacén del historial.
    """
    cache = _get_cached_historial()
    filas = cache.get("filas", [])

    codigo_buscado = str(codigo).strip().upper()
    cliente_buscado = str(cliente).strip()

    precio_cliente = None
    fecha_cliente = None
    precio_general = None
    fecha_general = None
    descripcion = ""
    almacen = ""

    for f in filas:
        if str(f.get("codigo") or "").strip().upper() != codigo_buscado:
            continue

        if not descripcion:
            descripcion = f.get("descripcion", "")
        if not almacen:
            almacen = f.get("almacen", "")

        precio = f.get("precio_unitario") or 0
        fecha = f.get("fecha_factura") or ""

        if cliente_buscado and cliente_buscado.lower() in (f.get("cliente") or "").lower():
            if fecha_cliente is None or (fecha and str(fecha) > str(fecha_cliente)):
                precio_cliente = precio
                fecha_cliente = fecha
        else:
            if fecha_general is None or (fecha and str(fecha) > str(fecha_general)):
                precio_general = precio
                fecha_general = fecha

    precio = precio_cliente if precio_cliente is not None else precio_general

    return {
        "codigo": codigo,
        "descripcion": descripcion,
        "almacen": almacen,
        "precio_unitario": precio,
        "fuente": "cliente" if precio_cliente is not None else ("general" if precio_general is not None else "ninguna"),
    }


@router.get("/vendedores")
def listar_vendedores(user: dict = Depends(get_current_user)):
    """Devuelve la lista de vendedores desde la hoja FIRMAS del Excel del cotizador."""
    settings = get_settings()
    excel_path = Path("Y:/COTIZACIONES/1. COTIZADOR/2. COTIZADOR 2.0.xlsm")
    vendedores = []
    if excel_path.exists():
        try:
            wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
            try:
                if "FIRMAS" in wb.sheetnames:
                    ws = wb["FIRMAS"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        nombre = str(row[0]).strip() if row and row[0] is not None else ""
                        if nombre and nombre.lower() not in ("none", "nan"):
                            vendedores.append(nombre)
            finally:
                wb.close()
        except Exception as exc:
            logger.warning("No se pudieron leer vendedores del Excel: %s", exc)

    # Siempre incluir al usuario actual como opción si no está en la lista
    usuario_nombre = user.get("nombre", "")
    if usuario_nombre and usuario_nombre not in vendedores:
        vendedores.insert(0, usuario_nombre)

    return {"vendedores": vendedores}


class LineaCotizacionInput(BaseModel):
    codigo: str
    descripcion: str
    almacen: str = ""
    cantidad: float = Field(default=1, ge=0)
    precio_unitario: float = Field(default=0, ge=0)
    descuento_pct: float = Field(default=0, ge=0, le=100)
    stock_leon: int = 0


class CotizacionInput(BaseModel):
    folio: str = ""
    cliente: str
    atencion: str = ""
    moneda: str = "USD"
    condiciones: str = "Contado"
    tiempo_entrega: str = "De 3-5 días después de su orden de compra y/o existencias en almacén y/o proveedor."
    leyenda_envio: str = ""
    con_descuento: bool = False
    con_stock_leon: bool = False
    vendedor: str = ""
    lineas: list[LineaCotizacionInput]


def _generar_folio(cliente: str, fecha: datetime.datetime | None = None) -> str:
    """Genera folio al estilo del Excel: CLIENTE YYMMDD."""
    if fecha is None:
        fecha = datetime.datetime.now()
    cliente_limpio = str(cliente).strip().upper()
    return f"{cliente_limpio} {fecha.strftime('%y%m%d')}"


@router.post("")
def guardar_cotizacion(
    data: CotizacionInput,
    user: dict = Depends(get_current_user),
):
    """Guarda una cotización y sus líneas."""
    logger.info("[guardar_cotizacion] Recibida peticion. Usuario=%s email=%s", user.get("nombre"), user.get("email"))
    logger.info("[guardar_cotizacion] Folio=%s Cliente=%s Moneda=%s Lineas=%d", data.folio, data.cliente, data.moneda, len(data.lineas))
    _init_cotizaciones_db()

    fecha = datetime.datetime.now()
    fecha_str = fecha.isoformat()
    folio = data.folio.strip() if data.folio.strip() else _generar_folio(data.cliente, fecha)

    # Calcular totales
    subtotal = 0.0
    lineas_procesadas = []
    for linea in data.lineas:
        precio_con_desc = linea.precio_unitario * (1 - linea.descuento_pct / 100)
        total_linea = linea.cantidad * precio_con_desc
        subtotal += total_linea
        lineas_procesadas.append({
            **linea.model_dump(),
            "precio_con_descuento": precio_con_desc,
            "total_linea": total_linea,
        })

    iva = subtotal * 0.16
    total = subtotal + iva

    conn = sqlite3.connect(str(_cotizaciones_db_path()))
    try:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO cotizaciones
            (folio, cliente, atencion, moneda, condiciones, tiempo_entrega, leyenda_envio,
             con_descuento, con_stock_leon, usuario_email, usuario_nombre, vendedor, subtotal, iva, total, fecha)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                folio,
                data.cliente.strip(),
                data.atencion.strip(),
                data.moneda.upper(),
                data.condiciones.strip(),
                data.tiempo_entrega.strip(),
                data.leyenda_envio.strip(),
                1 if data.con_descuento else 0,
                1 if data.con_stock_leon else 0,
                user.get("email", ""),
                user.get("nombre", ""),
                data.vendedor.strip(),
                subtotal,
                iva,
                total,
                fecha_str,
            ),
        )
        cotizacion_id = cur.lastrowid

        for linea in lineas_procesadas:
            cur.execute(
                """
                INSERT INTO cotizacion_lineas
                (cotizacion_id, codigo, descripcion, almacen, cantidad, precio_unitario,
                 descuento_pct, precio_con_descuento, total_linea, stock_leon)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    cotizacion_id,
                    linea["codigo"],
                    linea["descripcion"],
                    linea["almacen"],
                    linea["cantidad"],
                    linea["precio_unitario"],
                    linea["descuento_pct"],
                    linea["precio_con_descuento"],
                    linea["total_linea"],
                    int(linea.get("stock_leon", 0)),
                ),
            )

        conn.commit()
    finally:
        conn.close()

    return {
        "id": cotizacion_id,
        "folio": folio,
        "subtotal": subtotal,
        "iva": iva,
        "total": total,
    }


@router.get("")
def listar_cotizaciones(
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    busqueda: str = Query(""),
    user: dict = Depends(get_current_user),
):
    """Lista las cotizaciones guardadas."""
    _init_cotizaciones_db()
    conn = sqlite3.connect(str(_cotizaciones_db_path()))
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        sql = "SELECT * FROM cotizaciones WHERE 1=1"
        params = []
        if busqueda:
            sql += " AND (cliente LIKE ? OR folio LIKE ?)"
            params.extend([f"%{busqueda}%", f"%{busqueda}%"])
        sql += " ORDER BY fecha DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        cur.execute(sql, params)
        rows = cur.fetchall()
        return {"data": [dict(r) for r in rows]}
    finally:
        conn.close()


@router.get("/{cotizacion_id}")
def obtener_cotizacion(
    cotizacion_id: int,
    user: dict = Depends(get_current_user),
):
    """Obtiene una cotización con sus líneas."""
    _init_cotizaciones_db()
    conn = sqlite3.connect(str(_cotizaciones_db_path()))
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM cotizaciones WHERE id = ?", (cotizacion_id,))
        cot = cur.fetchone()
        if not cot:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        cur.execute("SELECT * FROM cotizacion_lineas WHERE cotizacion_id = ?", (cotizacion_id,))
        lineas = cur.fetchall()

        result = dict(cot)
        result["lineas"] = [dict(l) for l in lineas]
        return result
    finally:
        conn.close()


@router.get("/{cotizacion_id}/pdf")
def generar_pdf_cotizacion(
    cotizacion_id: int,
    user: dict = Depends(get_current_user),
):
    """Genera el PDF de una cotización con el formato corporativo de 3P."""
    _init_cotizaciones_db()
    conn = sqlite3.connect(str(_cotizaciones_db_path()))
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM cotizaciones WHERE id = ?", (cotizacion_id,))
        cot = cur.fetchone()
        if not cot:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        cur.execute("SELECT * FROM cotizacion_lineas WHERE cotizacion_id = ?", (cotizacion_id,))
        lineas = cur.fetchall()
    finally:
        conn.close()

    buffer = BytesIO()
    margin = 21.6  # 0.3 pulgadas en puntos
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=margin,
        leftMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
    )

    styles = getSampleStyleSheet()
    elements = []

    moneda = cot["moneda"]
    moneda_label = "USD" if moneda == "USD" else "MXN"
    ancho_util = letter[0] - 2 * margin

    # Fecha con formato largo en español
    try:
        fecha_dt = datetime.datetime.fromisoformat(cot["fecha"])
    except Exception:
        fecha_dt = datetime.datetime.now()
    meses = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
    ]
    dias = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
    fecha_larga = f"{dias[fecha_dt.weekday()]}, {fecha_dt.day} de {meses[fecha_dt.month - 1]} de {fecha_dt.year}"

    assets_dir = Path(__file__).parent / "assets"
    logo_path = assets_dir / "logo.png"

    # Encabezado con logo dentro de un recuadro
    logo_ancho = ancho_util - 36
    logo_alto = logo_ancho * 240 / 1600
    logo_img = Image(str(logo_path), width=logo_ancho, height=logo_alto)
    logo_img.hAlign = "CENTER"
    header_data = [[logo_img]]
    header_table = Table(header_data, colWidths=[ancho_util], rowHeights=[logo_alto + 16])
    header_table.setStyle(
        TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 1.5, colors.black),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ])
    )
    elements.append(header_table)
    elements.append(Spacer(1, 10))

    # Título
    elements.append(Paragraph("<font size='9'>FORMATO</font>", ParagraphStyle(name="Formato", alignment=1, fontSize=9)))
    elements.append(Paragraph("<font size='18'><b>C O T I Z A C I Ó N</b></font>", ParagraphStyle(name="Titulo", alignment=1, fontSize=18, spaceAfter=10)))

    # Folio y fecha (centrado: label + recuadro gris del folio, fecha a la derecha)
    fecha_style = ParagraphStyle(name="FechaLarga", fontSize=9, leading=11)
    folio_data = [
        ["", "FOLIO:", Paragraph(f"<b>{cot['folio']}</b>", styles["Normal"]), "", Paragraph(f"<b>{fecha_larga}</b>", fecha_style)],
    ]
    folio_table = Table(folio_data, colWidths=[ancho_util - 370, 45, 150, 15, 160])
    folio_table.setStyle(
        TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#E5E5E5")),
            ("BOX", (2, 0), (2, 0), 1, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
        ])
    )
    elements.append(folio_table)
    elements.append(Spacer(1, 12))

    # Cliente y Atención (cada uno en su propio recuadro gris)
    cliente_data = [
        ["Cliente:", Paragraph(f"<b>{cot['cliente']}</b>", styles["Normal"])],
        ["Atención a:", Paragraph(f"<b>{cot['atencion'] or ''}</b>", styles["Normal"])],
    ]
    cliente_table = Table(cliente_data, colWidths=[70, ancho_util - 70])
    cliente_table.setStyle(
        TableStyle([
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#E5E5E5")),
            ("BOX", (1, 0), (1, 0), 1, colors.black),
            ("BACKGROUND", (1, 1), (1, 1), colors.HexColor("#E5E5E5")),
            ("BOX", (1, 1), (1, 1), 1, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    elements.append(cliente_table)
    elements.append(Spacer(1, 14))

    # Presente
    elements.append(Paragraph("<font size='14'><b>P r e s e n t e</b></font>", ParagraphStyle(name="Presente", alignment=1, fontSize=14, spaceAfter=4)))
    elements.append(Paragraph("Sírvase encontrar a continuación la cotización que usted amablemente nos solicitó.", ParagraphStyle(name="TextoPresente", alignment=1, fontSize=10, spaceAfter=10)))

    # Tabla de líneas
    mostrar_stock_leon = bool(cot["con_stock_leon"])
    header = ["Código", "Cantidad", "Descripción", f"Precio Unitario {moneda_label}", f"TOTAL {moneda_label}"]
    if mostrar_stock_leon:
        header.insert(3, "Stock León")
    if cot["con_descuento"]:
        header.insert(5, "Desc %")

    table_data = [header]
    for l in lineas:
        row = [
            l["codigo"] or "—",
            f"{l['cantidad']:.0f}",
            Paragraph(l["descripcion"] or "—", styles["Normal"]),
        ]
        if mostrar_stock_leon:
            row.append(str(int(l["stock_leon"])))
        row.append(f"${l['precio_unitario']:,.2f}")
        if cot["con_descuento"]:
            row.append(f"{l['descuento_pct']:.0f}%")
        row.append(f"${l['total_linea']:,.2f}")
        table_data.append(row)

    # Ajustar anchos según combinación de columnas visibles
    if mostrar_stock_leon and cot["con_descuento"]:
        # Código, Cantidad, Descripción, Stock León, Precio Unitario, Desc %, Total
        col_widths = [70, 45, ancho_util - 475, 70, 80, 55, 80]
    elif mostrar_stock_leon:
        # Código, Cantidad, Descripción, Stock León, Precio Unitario, Total
        col_widths = [75, 50, ancho_util - 385, 70, 85, 90]
    elif cot["con_descuento"]:
        # Código, Cantidad, Descripción, Precio Unitario, Desc %, Total
        col_widths = [80, 55, ancho_util - 360, 90, 60, 90]
    else:
        # Código, Cantidad, Descripción, Precio Unitario, Total
        col_widths = [80, 55, ancho_util - 335, 90, 110]

    lineas_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    lineas_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.white),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 1.5, colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
        ])
    )
    elements.append(lineas_table)
    elements.append(Spacer(1, 10))

    # Totales
    totales_data = [
        ["", f"SUBTOTAL {moneda_label}", f"${cot['subtotal']:,.2f}"],
        ["", f"IVA 16%", f"${cot['iva']:,.2f}"],
        ["", f"TOTAL {moneda_label}", f"${cot['total']:,.2f}"],
    ]
    totales_table = Table(totales_data, colWidths=[ancho_util - 210, 110, 100])
    totales_table.setStyle(
        TableStyle([
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 2), (2, 2), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
        ])
    )
    elements.append(totales_table)
    elements.append(Spacer(1, 14))

    # Notas / condiciones
    notas_style = ParagraphStyle(name="Notas", fontSize=10, leading=14, spaceAfter=4)
    elements.append(Paragraph("<u>Incluye Asistencia Técnica</u>", notas_style))
    elements.append(Paragraph("<b>Estos precios son L.A.B. Su Granja</b>", notas_style))
    elements.append(Paragraph(f"<b>Plazo de Entrega:</b> {cot['tiempo_entrega']}", notas_style))
    elements.append(Paragraph(f"<b>Condiciones de Pago:</b> {cot['condiciones']}", notas_style))
    elements.append(Paragraph("<b>Vigencia de cotización:</b> 10 días.", notas_style))
    elements.append(Paragraph("<b>No incluye instalación mecánica, ni eléctrica.</b>", notas_style))
    if cot["leyenda_envio"]:
        elements.append(Paragraph(f"<b>Envío:</b> {cot['leyenda_envio']}", notas_style))
    if cot["con_descuento"]:
        elements.append(Paragraph("<b>Descuento aplicado por línea.</b>", notas_style))
    elements.append(Spacer(1, 20))

    # Firma
    firma_nombre = cot["vendedor"] or cot["usuario_nombre"] or "—"
    firma_path = _firma_path_para_vendedor(firma_nombre)

    firma_data = [
        [Paragraph("<b>Atentamente</b>", ParagraphStyle(name="Atentamente", alignment=1, fontSize=11))],
    ]
    if firma_path and firma_path.exists():
        firma_img = Image(str(firma_path), width=80, height=50)
        firma_img.hAlign = "CENTER"
        firma_data.append([firma_img])
    else:
        firma_data.append([Spacer(1, 36)])
    firma_data.append([Paragraph(f"<b>{firma_nombre}</b>", ParagraphStyle(name="NombreFirma", alignment=1, fontSize=10))])

    firma_table = Table(firma_data, colWidths=[ancho_util])
    firma_table.setStyle(
        TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E5E5E5")),
            ("BOX", (0, -1), (-1, -1), 1, colors.black),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    elements.append(firma_table)
    elements.append(Spacer(1, 14))

    # Pie de página
    footer_text = (
        "Industrial del Norte No. 201 Col. Industrial del Norte  TELS.: (477) 774-83-23 Y 774-83-26 C.P. 37200 León, Gto., México<br/>"
        "e-mail: ventas@3psadecv.com"
    )
    elements.append(Paragraph(footer_text, ParagraphStyle(name="Footer", alignment=1, fontSize=8, leading=12)))

    doc.build(elements)
    buffer.seek(0)

    from fastapi.responses import StreamingResponse

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={cot['folio']}.pdf"},
    )


def _firma_path_para_vendedor(nombre: str) -> Optional[Path]:
    assets_dir = Path(__file__).parent / "assets"
    mapping = {
        "america ruiz": assets_dir / "firma_america_ruiz.png",
        "carlos urbina": assets_dir / "firma_carlos_urbina.png",
        "cynthia hernandez": assets_dir / "firma_cynthia_hernandez.png",
    }
    clave = str(nombre).strip().lower()
    return mapping.get(clave)
