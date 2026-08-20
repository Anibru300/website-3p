import datetime
import mimetypes
import time
from pathlib import Path

from typing import List

from fastapi import APIRouter, Body, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook

from app.auth.dependencies import get_current_user
from app.config import get_settings
from app.database import postgres_cursor

router = APIRouter(prefix="/api/almacen", tags=["almacen"])


def _normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _read_excel_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [_normalize_text(cell.value) for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        item = {}
        for h, v in zip(headers, row):
            if h:
                item[h] = v
        data.append(item)
    return data


def _get_material_en_vales_by_code():
    """Lee el Excel de vales y devuelve un dict {codigo: cantidad_viva_total}.
    
    Si el archivo no existe o no se puede abrir, devuelve dict vacío.
    """
    settings = get_settings()
    excel_path = Path(settings.vales_excel_path)

    if not excel_path.exists():
        return {}

    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception:
        return {}

    try:
        cabeceras = _read_excel_sheet(wb, "VALES")
        detalles = _read_excel_sheet(wb, "DETALLE_VALES")
    finally:
        wb.close()

    cabeceras_by_folio = {}
    for c in cabeceras:
        folio = c.get("FOLIO_VALE")
        if folio is not None:
            cabeceras_by_folio[str(folio).strip()] = c

    material = {}
    for d in detalles:
        folio = str(d.get("FOLIO_VALE", "")).strip()
        cab = cabeceras_by_folio.get(folio, {})

        status_val = _normalize_text(cab.get("STATUS") or d.get("STATUS")).upper()
        cantidad_viva = d.get("CANTIDAD_VIVA", 0) or 0
        try:
            cantidad_viva = float(cantidad_viva)
        except (ValueError, TypeError):
            cantidad_viva = 0

        if status_val == "CERRADO" or cantidad_viva <= 0:
            continue

        codigo = _normalize_text(d.get("CODIGO"))
        if codigo:
            material[codigo] = material.get(codigo, 0) + cantidad_viva

    return material


def get_vales_abiertos_count():
    """Cuenta folios de vale distintos con cantidad viva > 0.

    Lee el Excel de vales directamente (la misma fuente que la sección de vales).
    """
    settings = get_settings()
    excel_path = Path(settings.vales_excel_path)

    if not excel_path.exists():
        return 0

    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception:
        return 0

    try:
        cabeceras = _read_excel_sheet(wb, "VALES")
        detalles = _read_excel_sheet(wb, "DETALLE_VALES")
    finally:
        wb.close()

    cabeceras_by_folio = {}
    for c in cabeceras:
        folio = c.get("FOLIO_VALE")
        if folio is not None:
            cabeceras_by_folio[str(folio).strip()] = c

    folios_abiertos = set()
    for d in detalles:
        folio = str(d.get("FOLIO_VALE", "")).strip()
        cab = cabeceras_by_folio.get(folio, {})

        status_val = _normalize_text(cab.get("STATUS") or d.get("STATUS")).upper()
        cantidad_viva = d.get("CANTIDAD_VIVA", 0) or 0
        try:
            cantidad_viva = float(cantidad_viva)
        except (ValueError, TypeError):
            cantidad_viva = 0

        if status_val == "CERRADO" or cantidad_viva <= 0:
            continue

        folios_abiertos.add(folio)

    return len(folios_abiertos)


@router.get("/existencias")
def existencias(
    limit: int = Query(50, ge=1, le=1000),
    offset: int = Query(0, ge=0, description="Registros a omitir para paginación"),
    busqueda: str = Query("", description="Filtrar por código o descripción"),
    almacen: str = Query("", description="Filtrar por clave de almacén (cve_alm)"),
    existencia: str = Query("con", description="Filtrar por existencia: con, sin, todos"),
    user: dict = Depends(get_current_user),
):
    # Material en vales desde el Excel de almacén (solo lectura)
    material_en_vales = _get_material_en_vales_by_code()

    almacen_filtro = ""
    params = {}
    if almacen:
        almacen_filtro = "AND cve_alm = %(almacen)s"
        params["almacen"] = almacen

    # Normalizar filtro de existencia
    existencia = (existencia or "con").lower().strip()
    if existencia not in ("con", "sin", "todos"):
        existencia = "con"

    existencia_filtro = ""
    if existencia == "con":
        existencia_filtro = "AND COALESCE(ef.existencia_total, 0) > 0"
    elif existencia == "sin":
        existencia_filtro = "AND COALESCE(ef.existencia_total, 0) = 0"
    # 'todos' no aplica filtro adicional

    base_sql = f"""
        WITH existencias_filtradas AS (
            SELECT cve_art, COALESCE(SUM(exist), 0) AS existencia_total
            FROM sae_existencias
            WHERE 1=1 {almacen_filtro}
            GROUP BY cve_art
        )
        SELECT
            sp.cve_art AS codigo,
            MAX(COALESCE(sp.descripcion, '')) AS descripcion,
            COALESCE(ef.existencia_total, 0) AS existencia_total
        FROM sae_productos sp
        LEFT JOIN existencias_filtradas ef ON ef.cve_art = sp.cve_art
        WHERE 1=1
    """
    count_sql = f"""
        WITH existencias_filtradas AS (
            SELECT cve_art, COALESCE(SUM(exist), 0) AS existencia_total
            FROM sae_existencias
            WHERE 1=1 {almacen_filtro}
            GROUP BY cve_art
        )
        SELECT COUNT(*) AS total
        FROM sae_productos sp
        LEFT JOIN existencias_filtradas ef ON ef.cve_art = sp.cve_art
        WHERE 1=1
    """
    if busqueda:
        filtro = """
            AND (
                LOWER(sp.cve_art) LIKE LOWER(%(busqueda)s)
                OR LOWER(COALESCE(sp.descripcion, '')) LIKE LOWER(%(busqueda)s)
            )
        """
        base_sql += filtro
        count_sql += filtro
        params["busqueda"] = f"%{busqueda}%"

    base_sql += f"""
        {existencia_filtro}
        GROUP BY sp.cve_art, ef.existencia_total
        ORDER BY MAX(COALESCE(sp.descripcion, ''))
        LIMIT %(limit)s OFFSET %(offset)s
    """
    count_sql += f" {existencia_filtro}"

    params["limit"] = limit
    params["offset"] = offset

    with postgres_cursor() as cur:
        cur.execute(base_sql, params)
        rows = cur.fetchall()

        cur.execute(count_sql, {k: v for k, v in params.items() if k in ("busqueda", "almacen")})
        total_row = cur.fetchone()
        total = total_row["total"] if total_row else 0

    data = []
    for row in rows:
        row_dict = dict(row)
        codigo = row_dict["codigo"]
        existencia_total = float(row_dict["existencia_total"] or 0)
        mat_vales = float(material_en_vales.get(codigo, 0))
        data.append({
            "codigo": codigo,
            "descripcion": row_dict["descripcion"],
            "existencia_total": existencia_total,
            "material_en_vales": mat_vales,
            "existencia_almacen": existencia_total - mat_vales,
        })

    return {"data": data, "total": total}


@router.post("/existencias-por-codigos")
def existencias_por_codigos(
    codigos: List[str] = Body(..., embed=True),
    user: dict = Depends(get_current_user),
):
    """Devuelve existencias totales y material en vales para una lista de códigos."""
    if not codigos:
        return {"data": {}}

    material_en_vales = _get_material_en_vales_by_code()

    sql = """
        SELECT cve_art AS codigo, COALESCE(SUM(exist), 0) AS existencia_total
        FROM sae_existencias
        WHERE cve_art = ANY(%(codigos)s)
        GROUP BY cve_art
    """

    with postgres_cursor() as cur:
        cur.execute(sql, {"codigos": list(set(c.strip() for c in codigos if c and str(c).strip()))})
        rows = cur.fetchall()

    data = {}
    for row in rows:
        row_dict = dict(row)
        codigo = row_dict["codigo"]
        existencia_total = float(row_dict["existencia_total"] or 0)
        mat_vales = float(material_en_vales.get(codigo, 0))
        data[codigo] = {
            "existencia_total": existencia_total,
            "material_en_vales": mat_vales,
            "existencia_almacen": existencia_total - mat_vales,
        }

    return {"data": data}


def _to_date(value):
    """Normaliza un valor de fecha de Excel a datetime.date."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


@router.get("/vales")
def vales(
    limit: int = Query(50, ge=1, le=500),
    busqueda: str = Query(""),
    responsable: str = Query("", description="Filtrar por responsable: joan, abelardo, aaron u otros"),
    fecha_desde: str = Query("", description="Fecha inicial YYYY-MM-DD"),
    fecha_hasta: str = Query("", description="Fecha final YYYY-MM-DD"),
    almacen: str = Query("", description="Filtrar por almacén origen"),
    user: dict = Depends(get_current_user),
):
    settings = get_settings()
    excel_path = Path(settings.vales_excel_path)

    if not excel_path.exists():
        raise HTTPException(
            status_code=503,
            detail=f"No se encontró el archivo de vales: {excel_path}",
        )

    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail=f"No se pudo abrir el archivo de vales. Puede estar abierto en Excel. Error: {exc}",
        )

    try:
        cabeceras = _read_excel_sheet(wb, "VALES")
        detalles = _read_excel_sheet(wb, "DETALLE_VALES")
    finally:
        wb.close()

    # Indexar cabeceras por folio
    cabeceras_by_folio = {}
    for c in cabeceras:
        folio = c.get("FOLIO_VALE")
        if folio is not None:
            cabeceras_by_folio[str(folio).strip()] = c

    # Combinar detalles con cabeceras
    resultados = []
    for d in detalles:
        folio = str(d.get("FOLIO_VALE", "")).strip()
        cab = cabeceras_by_folio.get(folio, {})

        # Filtrar solo vales abiertos / con cantidad viva
        status_val = _normalize_text(cab.get("STATUS") or d.get("STATUS")).upper()
        cantidad_viva = d.get("CANTIDAD_VIVA", 0) or 0
        try:
            cantidad_viva = float(cantidad_viva)
        except (ValueError, TypeError):
            cantidad_viva = 0

        if status_val == "CERRADO" or cantidad_viva <= 0:
            continue

        entregado_a = _normalize_text(cab.get("ENTREGADO_A"))

        # Filtro por responsable
        if responsable:
            responsable_lower = responsable.lower().strip()
            entregado_lower = entregado_a.lower()
            if responsable_lower == "otros":
                if any(n in entregado_lower for n in ["joan", "abelardo", "aaron"]):
                    continue
            elif responsable_lower not in entregado_lower:
                continue

        # Filtro por búsqueda
        if busqueda:
            busqueda_lower = busqueda.lower()
            campos = [
                str(folio),
                entregado_a,
                _normalize_text(d.get("CODIGO")),
                _normalize_text(d.get("DESCRIPCION")),
                _normalize_text(d.get("ALMACEN_ORIGEN")),
            ]
            if not any(busqueda_lower in c.lower() for c in campos if c):
                continue

        # Filtro por almacén origen
        almacen_origen = _normalize_text(d.get("ALMACEN_ORIGEN"))
        if almacen and almacen.lower() not in almacen_origen.lower():
            continue

        # Filtro por rango de fecha de salida
        fecha_salida = _to_date(cab.get("FECHA_SALIDA"))
        if fecha_salida and fecha_desde:
            try:
                desde = datetime.datetime.strptime(fecha_desde, "%Y-%m-%d").date()
                if fecha_salida < desde:
                    continue
            except ValueError:
                pass
        if fecha_salida and fecha_hasta:
            try:
                hasta = datetime.datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
                if fecha_salida > hasta:
                    continue
            except ValueError:
                pass

        resultados.append({
            "folio": folio,
            "entregado_a": entregado_a,
            "fecha_salida": cab.get("FECHA_SALIDA"),
            "codigo": _normalize_text(d.get("CODIGO")),
            "descripcion": _normalize_text(d.get("DESCRIPCION")),
            "cantidad": d.get("CANTIDAD", 0),
            "almacen_origen": _normalize_text(d.get("ALMACEN_ORIGEN")),
            "estado": _normalize_text(cab.get("STATUS") or d.get("STATUS")),
            "cantidad_viva": cantidad_viva,
        })

    # Ordenar por fecha de salida descendente
    resultados.sort(key=lambda x: x["fecha_salida"] or "", reverse=True)

    return {"data": resultados[:limit]}


@router.get("/subalmacenes")
def subalmacenes(user: dict = Depends(get_current_user)):
    sql = """
        SELECT
            e.cve_alm,
            MAX(a.descripcion) AS nombre,
            SUM(e.exist) AS existencia_total,
            SUM(e.exist * COALESCE(p.costo_promedio, 0)) AS valor_total
        FROM sae_existencias e
        LEFT JOIN sae_almacenes a ON e.cve_alm = a.cve_alm
        LEFT JOIN sae_productos p ON e.cve_art = p.cve_art
        WHERE e.exist > 0
        GROUP BY e.cve_alm
        ORDER BY valor_total DESC
    """

    with postgres_cursor() as cur:
        cur.execute(sql)
        rows = cur.fetchall()

    data = []
    for row in rows:
        row_dict = dict(row)
        data.append({
            "cve_alm": row_dict["cve_alm"],
            "nombre": row_dict["nombre"] or f"Almacén {row_dict['cve_alm']}",
            "existencia_total": float(row_dict["existencia_total"] or 0),
            "valor_total": float(row_dict["valor_total"] or 0),
        })

    return {"data": data}


# Caché simple del mapa codigo -> ruta_foto con TTL de 5 minutos
_FOTOS_CACHE = {"map": None, "ts": 0}
_FOTOS_TTL_SECONDS = 300


def _get_fotos_map():
    """Lee la hoja FOTOS_PRODUCTOS del Excel de almacén y devuelve un dict.

    Si el archivo no existe o no se puede abrir, devuelve dict vacío.
    La última versión del mapa se conserva en memoria por _FOTOS_TTL_SECONDS.
    """
    global _FOTOS_CACHE
    now = time.time()
    if _FOTOS_CACHE["map"] is not None and (now - _FOTOS_CACHE["ts"]) < _FOTOS_TTL_SECONDS:
        return _FOTOS_CACHE["map"]

    settings = get_settings()
    excel_path = Path(settings.vales_excel_path)

    if not excel_path.exists():
        _FOTOS_CACHE = {"map": {}, "ts": now}
        return {}

    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception:
        _FOTOS_CACHE = {"map": {}, "ts": now}
        return {}

    try:
        fotos = {}
        ws = wb["FOTOS_PRODUCTOS"]
        headers = [_normalize_text(cell.value) for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue
            item = {}
            for h, v in zip(headers, row):
                if h:
                    item[h] = v
            codigo = _normalize_text(item.get("CODIGO"))
            ruta = _normalize_text(item.get("RUTA_FOTO"))
            if codigo and ruta:
                fotos[codigo] = ruta
    finally:
        wb.close()

    _FOTOS_CACHE = {"map": fotos, "ts": now}
    return fotos


@router.get("/foto-producto/{codigo}")
def foto_producto(codigo: str, user: dict = Depends(get_current_user)):
    fotos = _get_fotos_map()
    ruta = fotos.get(codigo.strip())
    if not ruta:
        return Response(status_code=204)

    path = Path(ruta)
    if not path.exists():
        return Response(status_code=204)

    content_type, _ = mimetypes.guess_type(str(path))
    if not content_type:
        content_type = "image/jpeg"

    return StreamingResponse(open(path, "rb"), media_type=content_type)
