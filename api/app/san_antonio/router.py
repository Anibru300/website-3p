import os
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query
from openpyxl import load_workbook

from app.auth.dependencies import get_current_user
from app.config import get_settings

router = APIRouter(prefix="/api/san-antonio", tags=["san-antonio"])


def _normalize_key(key):
    if not key:
        return ""
    # Convertir a minúsculas, quitar espacios y caracteres especiales
    return str(key).strip().lower().replace(" ", "").replace("_", "")


def _normalize_row(row):
    return {
        _normalize_key(k): (v if v is not None else "")
        for k, v in row.items()
        if k
    }


@router.get("/ordenes")
def ordenes_san_antonio(
    limit: int = Query(50, ge=1, le=500),
    busqueda: str = Query(""),
    user: dict = Depends(get_current_user),
):
    settings = get_settings()
    excel_path = Path(settings.san_antonio_excel_path)

    if not excel_path.exists():
        raise HTTPException(
            status_code=503,
            detail=f"No se encontró el archivo de San Antonio: {excel_path}",
        )

    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail=f"No se pudo abrir el archivo de San Antonio. Puede estar abierto en Excel. Error: {exc}",
        )

    def read_sheet(sheet_name):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = ws.iter_rows(min_row=2, values_only=True)
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in rows:
            if not row or all(v is None for v in row):
                continue
            item = {}
            for h, v in zip(headers, row):
                if h:
                    item[h] = v
            data.append(item)
        return data

    try:
        cabeceras = read_sheet("OC_CABECERA")
        partidas = read_sheet("OC_PARTIDAS")
    finally:
        wb.close()

    # Normalizar claves a minúsculas y sin espacios/guiones bajos para el frontend
    cabeceras = [_normalize_row(c) for c in cabeceras]
    partidas = [_normalize_row(p) for p in partidas]

    busqueda_lower = busqueda.lower()
    if busqueda_lower:
        cabeceras = [
            c for c in cabeceras
            if any(busqueda_lower in str(v).lower() for v in c.values() if v is not None)
        ]
        folios_filtrados = {str(c.get("folio", "")).strip() for c in cabeceras}
        partidas = [
            p for p in partidas
            if str(p.get("folio", "")).strip() in folios_filtrados
        ]

    return {
        "total": len(cabeceras),
        "cabeceras": cabeceras[:limit],
        "partidas": partidas[:limit],
    }
