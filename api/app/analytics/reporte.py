"""Generación del reporte exportable de analytics (Excel y PDF).

Usa los mismos datos que el dashboard público (ver datos_publicos en
router.py) para que el reporte corresponda exactamente al período
visualizado por el usuario.
"""

import io
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

COLOR_ROJO = "C41E3A"
COLOR_GRIS = "6B7280"


def _etiqueta_periodo(dias: int, fecha_desde: Optional[str], fecha_hasta: Optional[str]) -> str:
    if fecha_desde or fecha_hasta:
        desde = fecha_desde or "inicio"
        hasta = fecha_hasta or "hoy"
        return f"{desde} a {hasta}"
    return f"Últimos {dias} días"


def _total(datos: list[dict]) -> int:
    return sum(int(d.get("total") or 0) for d in datos)


def _resolver_logo(logo_path: Optional[str]) -> Optional[Path]:
    if not logo_path:
        return None
    p = Path(logo_path)
    if not p.is_absolute():
        # Relativa al directorio api/ (este archivo: api/app/analytics/reporte.py)
        p = Path(__file__).resolve().parent.parent.parent / p
    if p.exists():
        return p
    logger.warning("[reporte] Logo no encontrado en %s; se omite", p)
    return None


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def generar_reporte_excel(
    datos: dict,
    dias: int,
    fecha_desde: Optional[str],
    fecha_hasta: Optional[str],
    logo_path: Optional[str] = None,
    generado_en: Optional[datetime] = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    generado_en = generado_en or datetime.now(timezone.utc)
    wb = Workbook()

    encabezado_font = Font(bold=True, color="FFFFFF", size=11)
    encabezado_fill = PatternFill("solid", fgColor=COLOR_ROJO)
    titulo_font = Font(bold=True, size=16, color="1F2937")
    subtitulo_font = Font(bold=True, size=12, color="1F2937")
    borde = Border(bottom=Side(style="thin", color="E5E7EB"))
    centro = Alignment(horizontal="center", vertical="center")

    def escribir_tabla(ws, fila_inicio: int, titulo: str, filas: list[dict], nombre_col: str = "Nombre"):
        ws.cell(row=fila_inicio, column=1, value=titulo).font = subtitulo_font
        fila = fila_inicio + 1
        ws.cell(row=fila, column=1, value=nombre_col).font = encabezado_font
        ws.cell(row=fila, column=1).fill = encabezado_fill
        ws.cell(row=fila, column=2, value="Eventos").font = encabezado_font
        ws.cell(row=fila, column=2).fill = encabezado_fill
        fila += 1
        if not filas:
            ws.cell(row=fila, column=1, value="Sin datos en el período").font = Font(italic=True, color=COLOR_GRIS)
            fila += 1
        else:
            for item in filas:
                ws.cell(row=fila, column=1, value=str(item.get("nombre") or item.get("dia") or item.get("hora") or ""))
                ws.cell(row=fila, column=2, value=int(item.get("total") or 0)).border = borde
                ws.cell(row=fila, column=1).border = borde
                fila += 1
        return fila + 1

    # --- Hoja Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    logo = _resolver_logo(logo_path)
    fila = 1
    if logo:
        try:
            img = XlImage(str(logo))
            max_w = 160
            if img.width > max_w:
                ratio = max_w / img.width
                img.width = max_w
                img.height = int(img.height * ratio)
            ws.add_image(img, "A1")
            fila = 10  # espacio para el logo
        except Exception as exc:
            logger.warning("[reporte] No se pudo incrustar el logo: %s", exc)

    ws.cell(row=fila, column=1, value="Reporte de tráfico del portal 3P").font = titulo_font
    ws.cell(row=fila + 1, column=1, value=f"Período: {_etiqueta_periodo(dias, fecha_desde, fecha_hasta)}").font = Font(size=11, color=COLOR_GRIS)
    ws.cell(row=fila + 2, column=1, value=f"Generado: {generado_en.astimezone().strftime('%d/%m/%Y %H:%M')}").font = Font(size=11, color=COLOR_GRIS)

    fila += 4
    totales = [
        ("Total de eventos", _total(datos["por_dia"])),
        ("Días con tráfico", len(datos["por_dia"])),
        ("Dispositivos distintos", len(datos["dispositivos"])),
        ("Países", len(datos["paises"])),
    ]
    for etiqueta, valor in totales:
        ws.cell(row=fila, column=1, value=etiqueta).font = Font(bold=True)
        ws.cell(row=fila, column=2, value=valor)
        fila += 1

    fila += 1
    fila = escribir_tabla(ws, fila, "Dispositivos", datos["dispositivos"])
    fila = escribir_tabla(ws, fila, "Navegadores", datos["navegadores"])
    fila = escribir_tabla(ws, fila, "Sistemas operativos", datos["sistemas"])
    fila = escribir_tabla(ws, fila, "Países", datos["paises"])
    fila = escribir_tabla(ws, fila, "Ciudades", datos["ciudades"])
    fila = escribir_tabla(ws, fila, "Origen del tráfico (referrers)", datos["referrers"])
    fila = escribir_tabla(ws, fila, "Páginas más visitadas", datos["paginas"], nombre_col="Página")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14

    # --- Hoja Visitas (por día y por hora, con gráficas) ---
    ws2 = wb.create_sheet("Visitas")
    ws2.sheet_view.showGridLines = False
    ws2.cell(row=1, column=1, value="Visitas por día").font = subtitulo_font
    fila = 3
    ws2.cell(row=fila, column=1, value="Fecha").font = encabezado_font
    ws2.cell(row=fila, column=1).fill = encabezado_fill
    ws2.cell(row=fila, column=2, value="Eventos").font = encabezado_font
    ws2.cell(row=fila, column=2).fill = encabezado_fill
    fila += 1
    inicio_dia = fila
    if datos["por_dia"]:
        for item in datos["por_dia"]:
            ws2.cell(row=fila, column=1, value=item["dia"])
            ws2.cell(row=fila, column=2, value=int(item["total"]))
            fila += 1
    else:
        ws2.cell(row=fila, column=1, value="Sin datos en el período").font = Font(italic=True, color=COLOR_GRIS)
        fila += 1
    fin_dia = fila - 1

    if datos["por_dia"] and fin_dia >= inicio_dia:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Visitas por día"
        chart.height = 8
        chart.width = 22
        data = Reference(ws2, min_col=2, min_row=inicio_dia - 1, max_row=fin_dia)
        cats = Reference(ws2, min_col=1, min_row=inicio_dia, max_row=fin_dia)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws2.add_chart(chart, "D3")

    fila += 2
    ws2.cell(row=fila, column=1, value="Visitas por hora (UTC)").font = subtitulo_font
    fila += 2
    ws2.cell(row=fila, column=1, value="Hora").font = encabezado_font
    ws2.cell(row=fila, column=1).fill = encabezado_fill
    ws2.cell(row=fila, column=2, value="Eventos").font = encabezado_font
    ws2.cell(row=fila, column=2).fill = encabezado_fill
    fila += 1
    inicio_hora = fila
    if datos["por_hora"]:
        for item in datos["por_hora"]:
            ws2.cell(row=fila, column=1, value=f"{item['hora']}:00")
            ws2.cell(row=fila, column=2, value=int(item["total"]))
            fila += 1
    else:
        ws2.cell(row=fila, column=1, value="Sin datos en el período").font = Font(italic=True, color=COLOR_GRIS)
        fila += 1
    fin_hora = fila - 1

    if datos["por_hora"] and fin_hora >= inicio_hora:
        chart2 = BarChart()
        chart2.type = "col"
        chart2.title = "Visitas por hora (UTC)"
        chart2.height = 8
        chart2.width = 22
        data2 = Reference(ws2, min_col=2, min_row=inicio_hora - 1, max_row=fin_hora)
        cats2 = Reference(ws2, min_col=1, min_row=inicio_hora, max_row=fin_hora)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws2.add_chart(chart2, f"D{fila + 2}")

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def generar_reporte_pdf(
    datos: dict,
    dias: int,
    fecha_desde: Optional[str],
    fecha_hasta: Optional[str],
    logo_path: Optional[str] = None,
    generado_en: Optional[datetime] = None,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Image,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    generado_en = generado_en or datetime.now(timezone.utc)
    rojo = colors.HexColor(f"#{COLOR_ROJO}")
    gris = colors.HexColor(f"#{COLOR_GRIS}")

    estilo_titulo = ParagraphStyle("titulo", fontName="Helvetica-Bold", fontSize=18, textColor=colors.HexColor("#1F2937"), spaceAfter=4)
    estilo_meta = ParagraphStyle("meta", fontName="Helvetica", fontSize=9.5, textColor=gris, spaceAfter=2)
    estilo_seccion = ParagraphStyle("seccion", fontName="Helvetica-Bold", fontSize=12, textColor=rojo, spaceBefore=14, spaceAfter=6)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
        title="Reporte de tráfico 3P",
    )

    story = []

    logo = _resolver_logo(logo_path)
    if logo:
        try:
            img = Image(str(logo))
            max_w = 4.5 * cm
            if img.drawWidth > max_w:
                ratio = max_w / img.drawWidth
                img.drawWidth = max_w
                img.drawHeight = img.drawHeight * ratio
            story.append(img)
            story.append(Spacer(1, 0.3 * cm))
        except Exception as exc:
            logger.warning("[reporte] No se pudo incrustar el logo en PDF: %s", exc)

    story.append(Paragraph("Reporte de tráfico del portal 3P", estilo_titulo))
    story.append(Paragraph(f"Período: {_etiqueta_periodo(dias, fecha_desde, fecha_hasta)}", estilo_meta))
    story.append(Paragraph(f"Generado: {generado_en.astimezone().strftime('%d/%m/%Y %H:%M')}", estilo_meta))
    story.append(Spacer(1, 0.3 * cm))

    # KPIs
    totales = [
        ["Total de eventos", str(_total(datos["por_dia"]))],
        ["Días con tráfico", str(len(datos["por_dia"]))],
        ["Dispositivos distintos", str(len(datos["dispositivos"]))],
        ["Países", str(len(datos["paises"]))],
    ]
    tabla_kpi = Table(totales, colWidths=[8 * cm, 4 * cm])
    tabla_kpi.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1F2937")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(tabla_kpi)

    def agregar_tabla(titulo: str, filas: list[dict], nombre_col: str = "Nombre", max_filas: int = 25):
        story.append(Paragraph(titulo, estilo_seccion))
        if not filas:
            story.append(Paragraph("<i>Sin datos en el período</i>", estilo_meta))
            return
        datos_tabla = [[nombre_col, "Eventos"]]
        for item in filas[:max_filas]:
            clave = item.get("nombre") or item.get("dia") or item.get("hora") or ""
            datos_tabla.append([str(clave), f"{int(item.get('total') or 0):,}"])
        tabla = Table(datos_tabla, colWidths=[12 * cm, 3 * cm], repeatRows=1)
        tabla.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BACKGROUND", (0, 0), (-1, 0), rojo),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, rojo),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(tabla)

    por_dia_con_fecha = [
        {**item, "nombre": item["dia"]} for item in datos["por_dia"]
    ]
    por_hora_etiquetada = [
        {**item, "nombre": f"{item['hora']}:00"} for item in datos["por_hora"]
    ]
    agregar_tabla("Visitas por día", por_dia_con_fecha, "Fecha", max_filas=40)
    agregar_tabla("Visitas por hora (UTC)", por_hora_etiquetada, "Hora")
    agregar_tabla("Dispositivos", datos["dispositivos"])
    agregar_tabla("Navegadores", datos["navegadores"])
    agregar_tabla("Sistemas operativos", datos["sistemas"])
    agregar_tabla("Países", datos["paises"])
    agregar_tabla("Ciudades", datos["ciudades"])
    agregar_tabla("Origen del tráfico (referrers)", datos["referrers"])
    agregar_tabla("Páginas más visitadas", datos["paginas"], "Página")

    doc.build(story)
    return buffer.getvalue()
