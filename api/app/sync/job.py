"""Job de sincronización Excel -> tablas sync_*.

SOLO LECTURA de los Excel: se copian a data/cache/ y se lee la copia con
openpyxl en modo read_only. Nunca se invoca save() ni se abre el original
en modo escritura (regla de oro, especialmente BD_ALMACEN_3P.xlsx).

Ejecutable:
    .venv/Scripts/python.exe -m app.sync.job
"""

import logging
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from app.config import get_settings
from app.database import users_connection
from app.services.excel import read_excel_sheet, normalize_text
from app.services.fuentes import FUENTES_EXCEL
from app.sync import db as sync_db

logger = logging.getLogger(__name__)

# Fase 2: solo E1 (vales) y E2 (pedidos). E3/E4/E5 quedan fuera por ahora.
FUENTES_HABILITADAS = ("vales", "pedidos")


def _ahora() -> str:
    return datetime.now(timezone.utc).isoformat()


def _headers_de_hoja(ws, header_row: int) -> list[str]:
    fila = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    if not fila:
        return []
    return [normalize_text(v) for v in fila[0]]


def _leer_sheets_de_copia(ruta_copia: Path, spec: dict) -> dict[str, list[dict]]:
    """Abre la copia local (read_only) y valida hojas/columnas esperadas."""
    wb = load_workbook(filename=str(ruta_copia), read_only=True, data_only=True)
    try:
        header_row = spec.get("header_row", 1)
        sheets = {}
        for sheet_name, columnas_esperadas in spec["sheets"].items():
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo")
            ws = wb[sheet_name]
            headers = set(_headers_de_hoja(ws, header_row))
            faltantes = [c for c in columnas_esperadas if c not in headers]
            if faltantes:
                raise ValueError(
                    f"Hoja '{sheet_name}': columnas faltantes {faltantes}"
                )
            sheets[sheet_name] = read_excel_sheet(wb, sheet_name)
        return sheets
    finally:
        wb.close()


def sincronizar_fuente(fuente: str) -> dict:
    """Sync de una fuente habilitada. Devuelve resumen y nunca lanza excepción."""
    if fuente not in FUENTES_HABILITADAS:
        raise ValueError(f"Fuente no habilitada para sync: {fuente}")

    settings = get_settings()
    spec = FUENTES_EXCEL[fuente]
    ruta = Path(getattr(settings, spec["setting"]))
    inicio = _ahora()
    resumen = {"fuente": fuente, "estado": "error", "filas": 0, "error": None}

    with users_connection() as conn:
        sync_db.init_sync_db(conn)
        try:
            if not ruta.exists():
                raise FileNotFoundError(f"No existe o no accesible: {ruta}")

            mtime = datetime.fromtimestamp(
                ruta.stat().st_mtime, tz=timezone.utc
            ).isoformat()

            # Copia local: evita locks del archivo abierto en red (R4).
            cache_dir = Path(settings.users_db_path).resolve().parent / "cache"
            cache_dir.mkdir(parents=True, exist_ok=True)
            copia = cache_dir / f"sync_{fuente}{ruta.suffix}"
            shutil.copy2(ruta, copia)

            sheets = _leer_sheets_de_copia(copia, spec)
            total = sync_db.guardar_sheets(conn, fuente, sheets, mtime)
            sync_db.registrar_sync(conn, fuente, inicio, "ok", filas=total, mtime=mtime)
            conn.commit()
            resumen.update({"estado": "ok", "filas": total, "mtime": mtime})
            logger.info("[sync] %s: ok (%d filas)", fuente, total)
        except Exception as exc:  # noqa: BLE001
            conn.rollback()
            sync_db.registrar_sync(conn, fuente, inicio, "error", error=str(exc))
            conn.commit()
            resumen["error"] = str(exc)
            logger.warning("[sync] %s: error: %s", fuente, exc)
    return resumen


def sincronizar_todas() -> list[dict]:
    resultados = [sincronizar_fuente(f) for f in FUENTES_HABILITADAS]
    # Mantener la demanda por pedido al día con cada sync (mejor esfuerzo).
    try:
        from app.services.logistica import regenerar_demanda

        resumen = regenerar_demanda()
        logger.info("[sync] demanda logística regenerada: %s", resumen)
    except Exception as exc:  # noqa: BLE001
        logger.warning("[sync] No se pudo regenerar la demanda logística: %s", exc)
    return resultados


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    resultados = sincronizar_todas()
    for r in resultados:
        if r["estado"] == "ok":
            print(f"[OK] {r['fuente']}: {r['filas']} filas")
        else:
            print(f"[ERROR] {r['fuente']}: {r['error']}")
    sys.exit(0 if all(r["estado"] == "ok" for r in resultados) else 1)
