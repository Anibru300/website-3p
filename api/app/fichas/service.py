"""Servicios del módulo Fichas técnicas / documentos de producto.

Gestiona el registro en SQLite (data/fichas.db) y los archivos PDF en disco
(data/fichas/<marca>/<codigo>/<tipo>/<timestamp>_<hex>.pdf). Cada subida crea
una NUEVA versión del documento: la fila vigente anterior pasa a historial
(vigente=0) sin borrar fila ni archivo. Los productos disponibles salen del
Excel de almacén (BD_ALMACEN_3P.xlsx, SOLO LECTURA): cada marca tiene una hoja
con encabezados en la fila 4 (CODIGO, DESCRIPCION, ...) y datos de la fila 5
en adelante. No se permiten documentos huérfanos: el código debe existir en
la hoja de su marca.
"""

import re
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from app.config import get_settings
from app.database import fichas_connection
from app.services.excel import normalize_text

# Tamaño máximo de PDF aceptado: 25 MB.
MAX_PDF_BYTES = 25 * 1024 * 1024

# Mapeo slug de marca (usado en el sitio público) -> hoja del Excel de almacén.
# Las hojas son estables (una por almacén/marca); no se parametrizan.
MARCAS = {
    "fancom": {"nombre": "FANCOM", "hoja": "40 A-40 FANCOM"},
    "lubing": {"nombre": "LUBING", "hoja": "2 A-2 LUBING"},
    "sbm": {"nombre": "SBM", "hoja": "20 A-20 SBM"},
    "lbwhite": {"nombre": "LB White", "hoja": "26 A-26 L.B. WHITE"},
    "georgia-poultry": {"nombre": "GEORGIA POULTRY", "hoja": "10 A-10 GEORGIA POULTRY"},
    "amt": {"nombre": "AMT", "hoja": "45 A-45 AMT INC"},
    "alke": {"nombre": "ALKE", "hoja": "48 A-48 ALKE"},
    "ms-schippers": {"nombre": "MS Schippers", "hoja": "41 A-41 MS DETERGENTES"},
}


class MarcaNoSoportada(ValueError):
    pass


class ExcelNoDisponible(RuntimeError):
    pass


class ProductoNoEncontrado(LookupError):
    """El código no existe en la hoja de la marca del Excel de almacén."""


class TipoDocumentoNoValido(ValueError):
    pass


def listar_tipos() -> list[dict]:
    """Tipos de documento activos, en el orden del catálogo."""
    with fichas_connection() as conn:
        rows = conn.execute(
            """
            SELECT codigo, nombre, orden, activo
            FROM tipos_documento
            WHERE activo = 1
            ORDER BY orden
            """
        ).fetchall()
    return [dict(r) for r in rows]


def listar_marcas() -> list[dict]:
    return [{"slug": slug, "nombre": info["nombre"]} for slug, info in MARCAS.items()]


def _fichas_dir() -> Path:
    return Path(get_settings().fichas_dir)


def _normalizar_codigo(codigo: str) -> str:
    return normalize_text(codigo).upper()


def _ahora() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _codigo_seguro(codigo: str) -> str:
    """Segmento de carpeta derivado del código: sin caracteres de riesgo."""
    return re.sub(r"[^\w\-.]+", "_", codigo.strip()) or "sin_codigo"


def _incrementar_version(version: str) -> str:
    """Auto-incrementa la versión: '2'->'3', '1.0'->'1.1' (sube el último segmento)."""
    partes = version.split(".")
    try:
        partes[-1] = str(int(partes[-1]) + 1)
    except ValueError:
        return version
    return ".".join(partes)


# ---------------------------------------------------------------------------
# Productos por marca (lectura del Excel de almacén, SOLO LECTURA)
# ---------------------------------------------------------------------------


def productos_de_marca(slug: str) -> list[dict]:
    """Productos de una marca desde su hoja del Excel: [{codigo, descripcion}].

    Raises:
        MarcaNoSoportada: si el slug no está en el catálogo.
        ExcelNoDisponible: si el Excel no existe, no se puede abrir o le falta
            la hoja de la marca.
    """
    info = MARCAS.get(slug)
    if info is None:
        raise MarcaNoSoportada(f"Marca no soportada: {slug}")

    excel_path = Path(get_settings().vales_excel_path)
    if not excel_path.exists():
        raise ExcelNoDisponible(f"No se encontró el Excel de almacén: {excel_path}")

    try:
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ExcelNoDisponible(f"No se pudo abrir el Excel de almacén: {exc}") from exc

    try:
        if info["hoja"] not in wb.sheetnames:
            raise ExcelNoDisponible(f"La hoja '{info['hoja']}' no existe en el Excel de almacén")
        ws = wb[info["hoja"]]
        filas = ws.iter_rows(min_row=4, values_only=True)
        encabezados = [normalize_text(v) for v in next(filas, ())]
        try:
            idx_codigo = encabezados.index("CODIGO")
            idx_descripcion = encabezados.index("DESCRIPCION")
        except ValueError as exc:
            raise ExcelNoDisponible(
                f"La hoja '{info['hoja']}' no tiene las columnas CODIGO/DESCRIPCION"
            ) from exc

        productos = []
        for fila in filas:
            codigo = normalize_text(fila[idx_codigo] if idx_codigo < len(fila) else None)
            if not codigo:
                continue
            productos.append(
                {
                    "codigo": codigo,
                    "descripcion": normalize_text(
                        fila[idx_descripcion] if idx_descripcion < len(fila) else None
                    ),
                }
            )
        return productos
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Documentos de producto (SQLite + archivos en disco)
# ---------------------------------------------------------------------------


def _fila_a_documento(row) -> dict:
    doc = dict(row)
    doc["url"] = f"/api/fichas/{doc['id']}/pdf"
    return doc


def listar_documentos(
    marca: str | None = None,
    codigo: str | None = None,
    q: str | None = None,
    tipo: str | None = None,
    publico: int | None = None,
    vigente: int | None = None,
    activo: int | None = None,
) -> list[dict]:
    sql = (
        "SELECT d.*, t.codigo AS tipo, t.nombre AS tipo_nombre "
        "FROM documentos_producto d "
        "JOIN tipos_documento t ON t.id = d.tipo_documento_id "
        "WHERE 1=1"
    )
    params: list = []
    if marca:
        sql += " AND d.marca = ?"
        params.append(marca)
    if codigo:
        sql += " AND d.codigo = ?"
        params.append(_normalizar_codigo(codigo))
    if q:
        sql += (
            " AND (d.codigo LIKE ? OR d.descripcion LIKE ?"
            " OR d.nombre_documento LIKE ? OR d.nombre_archivo LIKE ?)"
        )
        like = f"%{q.strip()}%"
        params.extend([like, like, like, like])
    if tipo:
        sql += " AND t.codigo = ?"
        params.append(tipo)
    if publico is not None:
        sql += " AND d.publico = ?"
        params.append(int(publico))
    if vigente is not None:
        sql += " AND d.vigente = ?"
        params.append(int(vigente))
    if activo is not None:
        sql += " AND d.activo = ?"
        params.append(int(activo))
    sql += " ORDER BY d.marca, d.codigo, t.orden, d.fecha_carga DESC"
    with fichas_connection() as conn:
        return [_fila_a_documento(r) for r in conn.execute(sql, params).fetchall()]


def obtener_documento(doc_id: int) -> dict | None:
    with fichas_connection() as conn:
        row = conn.execute(
            """
            SELECT d.*, t.codigo AS tipo, t.nombre AS tipo_nombre
            FROM documentos_producto d
            JOIN tipos_documento t ON t.id = d.tipo_documento_id
            WHERE d.id = ?
            """,
            (doc_id,),
        ).fetchone()
    return _fila_a_documento(row) if row else None


def ruta_archivo(doc: dict) -> Path:
    """Ruta absoluta del PDF en disco, verificando anti path traversal."""
    base = _fichas_dir().resolve()
    ruta = (base / doc["archivo"]).resolve()
    if not ruta.is_relative_to(base):
        raise ValueError(f"Ruta de archivo fuera del directorio permitido: {doc['archivo']}")
    return ruta


def guardar_documento(
    marca: str,
    codigo: str,
    tipo: str,
    contenido: bytes,
    nombre_original: str,
    usuario: str,
    nombre_documento: str = "",
    descripcion_documento: str = "",
    numero_documento: str = "",
    version: str = "",
    fecha_documento: str = "",
    publico: int = 1,
) -> dict:
    """Sube una nueva versión del documento (marca, codigo, tipo).

    La fila vigente anterior del mismo tipo pasa a historial (vigente=0, sin
    borrar fila ni archivo). El código debe existir en la hoja de la marca del
    Excel de almacén (no se permiten documentos huérfanos).

    Raises:
        MarcaNoSoportada: marca fuera del catálogo.
        ProductoNoEncontrado: el código no existe en la hoja de la marca.
        ExcelNoDisponible: el Excel de almacén no se pudo leer.
        TipoDocumentoNoValido: el tipo de documento no existe.
        ValueError: PDF vacío o que supera MAX_PDF_BYTES.
    """
    if marca not in MARCAS:
        raise MarcaNoSoportada(f"Marca no soportada: {marca}")
    codigo = _normalizar_codigo(codigo)
    if not codigo:
        raise ValueError("El código de producto es obligatorio")
    if not contenido:
        raise ValueError("El PDF está vacío")
    if len(contenido) > MAX_PDF_BYTES:
        raise ValueError("El PDF supera el tamaño máximo permitido (25 MB)")

    productos = productos_de_marca(marca)  # valida contra el Excel (puede lanzar ExcelNoDisponible)
    producto = next((p for p in productos if p["codigo"] == codigo), None)
    if producto is None:
        raise ProductoNoEncontrado(f"El código '{codigo}' no existe en la marca '{marca}'")
    descripcion = producto["descripcion"]

    with fichas_connection() as conn:
        tipo_row = conn.execute(
            "SELECT id FROM tipos_documento WHERE codigo = ? AND activo = 1", (tipo,)
        ).fetchone()
        if tipo_row is None:
            raise TipoDocumentoNoValido(f"Tipo de documento no válido: {tipo}")
        tipo_id = tipo_row["id"]

        anterior = conn.execute(
            """
            SELECT id, version FROM documentos_producto
            WHERE marca = ? AND codigo = ? AND tipo_documento_id = ? AND vigente = 1 AND activo = 1
            ORDER BY fecha_carga DESC
            LIMIT 1
            """,
            (marca, codigo, tipo_id),
        ).fetchone()

        if not version:
            version = _incrementar_version(anterior["version"]) if anterior else "1.0"

        # Nuevo archivo en disco: nunca se sobrescribe, cada versión es un archivo nuevo.
        nombre_seguro = (
            f"{datetime.now().strftime('%Y%m%d-%H%M%S')}_{uuid.uuid4().hex[:8]}.pdf"
        )
        rel_path = Path(marca) / _codigo_seguro(codigo) / tipo / nombre_seguro
        abs_path = _fichas_dir() / rel_path
        abs_path.parent.mkdir(parents=True, exist_ok=True)
        abs_path.write_bytes(contenido)

        ahora = _ahora()
        if anterior:
            conn.execute(
                """
                UPDATE documentos_producto
                SET vigente = 0, fecha_modificacion = ?, usuario_modificacion = ?
                WHERE id = ?
                """,
                (ahora, usuario, anterior["id"]),
            )

        cur = conn.execute(
            """
            INSERT INTO documentos_producto (
                marca, codigo, descripcion, tipo_documento_id, nombre_documento,
                descripcion_documento, numero_documento, version, nombre_archivo,
                archivo, tamano, mime_type, fecha_documento, publico, vigente,
                activo, fecha_carga, usuario_carga
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'application/pdf', ?, ?, 1, 1, ?, ?)
            """,
            (
                marca,
                codigo,
                descripcion,
                tipo_id,
                nombre_documento.strip(),
                descripcion_documento.strip(),
                numero_documento.strip(),
                version.strip(),
                nombre_original,
                str(rel_path),
                len(contenido),
                fecha_documento.strip(),
                int(publico),
                ahora,
                usuario,
            ),
        )
        doc_id = cur.lastrowid
        conn.commit()
    return obtener_documento(doc_id)


CAMPOS_EDITABLES = (
    "nombre_documento",
    "descripcion_documento",
    "numero_documento",
    "version",
    "fecha_documento",
    "publico",
    "vigente",
    "activo",
)


def actualizar_documento(doc_id: int, cambios: dict, usuario: str) -> dict | None:
    """Actualiza los campos editables de un documento. None si no existe."""
    campos = {k: v for k, v in cambios.items() if k in CAMPOS_EDITABLES and v is not None}
    if not campos:
        return obtener_documento(doc_id)
    asignaciones = ", ".join(f"{k} = ?" for k in campos)
    valores = [int(v) if k in ("publico", "vigente", "activo") else str(v) for k, v in campos.items()]
    with fichas_connection() as conn:
        existe = conn.execute(
            "SELECT id FROM documentos_producto WHERE id = ?", (doc_id,)
        ).fetchone()
        if existe is None:
            return None
        conn.execute(
            f"""
            UPDATE documentos_producto
            SET {asignaciones}, fecha_modificacion = ?, usuario_modificacion = ?
            WHERE id = ?
            """,
            (*valores, _ahora(), usuario, doc_id),
        )
        conn.commit()
    return obtener_documento(doc_id)


def eliminar_documento(doc_id: int, usuario: str) -> bool:
    """Eliminación LÓGICA (activo=0). No borra la fila ni el archivo físico."""
    with fichas_connection() as conn:
        existe = conn.execute(
            "SELECT id FROM documentos_producto WHERE id = ?", (doc_id,)
        ).fetchone()
        if existe is None:
            return False
        conn.execute(
            """
            UPDATE documentos_producto
            SET activo = 0, fecha_modificacion = ?, usuario_modificacion = ?
            WHERE id = ?
            """,
            (_ahora(), usuario, doc_id),
        )
        conn.commit()
    return True


def listar_publicas(marca: str, codigo: str) -> list[dict]:
    """Lista mínima para el sitio público: solo públicas, vigentes y activas."""
    docs = listar_documentos(
        marca=marca, codigo=codigo, publico=1, vigente=1, activo=1
    )
    return [
        {
            "id": d["id"],
            "marca": d["marca"],
            "codigo": d["codigo"],
            "tipo": d["tipo"],
            "tipo_nombre": d["tipo_nombre"],
            "nombre_documento": d["nombre_documento"],
            "version": d["version"],
            "fecha_documento": d["fecha_documento"],
            "url": d["url"],
        }
        for d in docs
    ]


def visible_publicamente(doc: dict) -> bool:
    return bool(doc["publico"] and doc["vigente"] and doc["activo"])
