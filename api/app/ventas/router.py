import logging
from datetime import date
from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.auth.dependencies import get_current_user
from app.database import postgres_cursor
from app.services.excel import (
    _get_cached_historial,
    get_pedido_detalle_excel,
    get_pedidos_vivos_excel,
    precargar_historial_cache,
)

router = APIRouter(prefix="/api/ventas", tags=["ventas"])
logger = logging.getLogger(__name__)


@router.get("/pedidos-vivos")
def pedidos_vivos(
    limit: int = Query(50, ge=1, le=500),
    busqueda: str = Query(""),
    user: dict = Depends(get_current_user),
):
    resultados = get_pedidos_vivos_excel(busqueda=busqueda, limit=limit)
    return {"data": resultados}


@router.get("/facturas-cobranza")
def facturas_cobranza(
    limit: int = Query(50, ge=1, le=500),
    user: dict = Depends(get_current_user),
):
    sql = """
        SELECT
            cve_doc AS folio,
            cliente,
            fecha_doc,
            importe_total AS total,
            estado_cobranza
        FROM v_facturas_cobranza
        WHERE estado_cobranza IN ('Pendiente', 'Vencida')
        ORDER BY fecha_doc DESC
        LIMIT %(limit)s
    """
    with postgres_cursor() as cur:
        cur.execute(sql, {"limit": limit})
        rows = cur.fetchall()

    return {"data": [dict(row) for row in rows]}


@router.get("/seguimiento-documental")
def seguimiento_documental(
    folio_pedido: str = Query(""),
    limit: int = Query(50, ge=1, le=500),
    user: dict = Depends(get_current_user),
):
    try:
        sql = """
            SELECT
                folio_pedido,
                fecha_pedido,
                cliente,
                codigo,
                descripcion,
                cantidad_pedido,
                folio_remision,
                cantidad_remision,
                folio_factura,
                cantidad_factura,
                estatus_linea
            FROM v_seguimiento_documental
            WHERE 1=1
        """
        params = {}
        if folio_pedido:
            sql += " AND folio_pedido = %(folio_pedido)s"
            params["folio_pedido"] = folio_pedido
        sql += " ORDER BY folio_pedido, codigo LIMIT %(limit)s"
        params["limit"] = limit

        with postgres_cursor() as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

        return {"data": [dict(row) for row in rows]}
    except Exception as e:
        logger.warning("v_seguimiento_documental fallo, usando fallback Excel: %s", e)
        try:
            detalle = get_pedido_detalle_excel(folio_pedido)
            return {"data": detalle}
        except Exception as e2:
            logger.error("Fallback Excel tambien fallo: %s", e2)
            return {"data": []}


def _parse_iso_date(value):
    """Convierte un string ISO (YYYY-MM-DD) o datetime a date; devuelve None si no es posible."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        try:
            return date.fromisoformat(value[:10])
        except ValueError:
            return None
    if hasattr(value, "date"):
        # datetime.datetime
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _filtrar_historial(
    cache_filas,
    busqueda: str = "",
    cliente: Optional[List[str]] = None,
    codigo: Optional[List[str]] = None,
    moneda: str = "",
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
):
    """Aplica filtros a la lista de filas del historial de ventas."""
    busqueda_lower = busqueda.lower().strip()
    clientes_lower = [c.lower().strip() for c in (cliente or []) if c and c.strip()]
    codigos_lower = [c.lower().strip() for c in (codigo or []) if c and c.strip()]
    moneda_upper = moneda.upper().strip()

    resultados = []
    for fila in cache_filas:
        # Filtro por búsqueda general
        if busqueda_lower and not any(
            busqueda_lower in campo
            for campo in (fila["cliente"].lower(), fila["codigo"].lower(), fila["descripcion"].lower())
        ):
            continue

        # Filtros específicos
        if clientes_lower and not any(c in fila["cliente"].lower() for c in clientes_lower):
            continue
        if codigos_lower and not any(c in fila["codigo"].lower() for c in codigos_lower):
            continue
        if moneda_upper and moneda_upper not in fila["moneda"]:
            continue

        # Filtro por fecha de factura
        fila_fecha = _parse_iso_date(fila.get("fecha_factura"))
        if fila_fecha:
            if fecha_desde and fila_fecha < fecha_desde:
                continue
            if fecha_hasta and fila_fecha > fecha_hasta:
                continue

        resultados.append(fila)

    return resultados


@router.get("/historial")
def historial_ventas(
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    busqueda: str = Query(""),
    cliente: List[str] = Query(default_factory=list),
    codigo: List[str] = Query(default_factory=list),
    moneda: str = Query(""),
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    user: dict = Depends(get_current_user),
):
    """Lee el historial de ventas desde la caché en memoria.

    Filtra solo filas donde 'Tipo de Fila' contenga la palabra 'factura'.
    Permite buscar por cliente, código, descripción, moneda o rango de fecha.
    Soporta paginación con limit/offset.
    """
    cache = _get_cached_historial()
    if not cache["filas"]:
        from app.config import get_settings
        from pathlib import Path

        settings = get_settings()
        excel_path = Path(settings.ventas_facturacion_excel_path)
        if not excel_path.exists():
            raise HTTPException(
                status_code=503,
                detail=f"No se encontró el archivo de ventas: {excel_path}",
            )
        raise HTTPException(
            status_code=503,
            detail="No se pudo cargar el historial de ventas. Puede estar abierto en Excel.",
        )

    resultados = _filtrar_historial(
        cache["filas"],
        busqueda=busqueda,
        cliente=cliente,
        codigo=codigo,
        moneda=moneda,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )

    total = len(resultados)
    totales = {"MXN": 0.0, "USD": 0.0}
    for r in resultados:
        moneda = r["moneda"] if r["moneda"] == "USD" else "MXN"
        totales[moneda] += r["importe_partida"]

    return {
        "data": resultados[offset : offset + limit],
        "total": total,
        "totales": totales,
        "limit": limit,
        "offset": offset,
    }


class HistorialExportRequest(BaseModel):
    busqueda: str = ""
    cliente: List[str] = []
    codigo: List[str] = []
    moneda: str = ""
    fecha_desde: Optional[str] = None
    fecha_hasta: Optional[str] = None


def _generar_excel_historial(filas):
    """Genera un archivo Excel formateado con el historial de ventas filtrado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de ventas"

    # Encabezados
    headers = [
        "Fecha factura",
        "Fecha pedido",
        "Cliente",
        "Código",
        "Descripción",
        "Cantidad",
        "Precio unitario",
        "Importe partida",
        "Moneda",
        "Tipo de cambio",
        "Almacén",
        "Folio factura",
        "Folio pedido",
    ]

    # Estilos
    header_fill = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    money_fmt = '#,##0.00'
    date_fmt = 'DD/MM/YYYY'

    # Escribir encabezados
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Congelar primera fila
    ws.freeze_panes = "A2"

    # Escribir datos
    row_idx = 2
    totales_mxn = 0.0
    totales_usd = 0.0
    for fila in filas:
        moneda = fila.get("moneda", "MXN").upper() or "MXN"
        importe = fila.get("importe_partida", 0.0) or 0.0
        if moneda == "USD":
            totales_usd += importe
        else:
            totales_mxn += importe

        values = [
            _parse_iso_date(fila.get("fecha_factura")),
            _parse_iso_date(fila.get("fecha_pedido")),
            fila.get("cliente", ""),
            fila.get("codigo", ""),
            fila.get("descripcion", ""),
            fila.get("cantidad", 0),
            fila.get("precio_unitario", 0.0),
            importe,
            moneda,
            fila.get("tipo_cambio", 0.0),
            fila.get("almacen", ""),
            fila.get("folio_factura", ""),
            fila.get("folio_pedido", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Formato de moneda
            if col_idx in (7, 8):  # precio_unitario, importe_partida
                cell.number_format = money_fmt
            elif col_idx == 10:  # tipo_cambio
                cell.number_format = '0.0000'
            elif col_idx in (1, 2):  # fechas
                cell.number_format = date_fmt

            # Color de fondo por moneda
            if col_idx == 9 and value == "USD":
                cell.fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")

        row_idx += 1

    # Fila de totales
    total_row = row_idx + 1
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=totales_mxn).number_format = money_fmt
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value="MXN").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=8, value=totales_usd).number_format = money_fmt
    ws.cell(row=total_row + 1, column=8).font = Font(bold=True)
    ws.cell(row=total_row + 1, column=9, value="USD").font = Font(bold=True)

    # Ancho de columnas
    col_widths = [14, 14, 35, 18, 45, 12, 16, 18, 10, 14, 25, 18, 18]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Altura de encabezado
    ws.row_dimensions[1].height = 24

    # Autofiltro
    end_col = get_column_letter(len(headers))
    end_row = max(row_idx - 1, 1)
    ws.auto_filter.ref = f"A1:{end_col}{end_row}"

    # Guardar a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.post("/historial/exportar")
def exportar_historial_ventas(
    request: HistorialExportRequest,
    user: dict = Depends(get_current_user),
):
    """Exporta el historial de ventas filtrado a un archivo Excel formateado."""
    cache = _get_cached_historial()
    if not cache["filas"]:
        from app.config import get_settings
        from pathlib import Path

        settings = get_settings()
        excel_path = Path(settings.ventas_facturacion_excel_path)
        if not excel_path.exists():
            raise HTTPException(
                status_code=503,
                detail=f"No se encontró el archivo de ventas: {excel_path}",
            )
        raise HTTPException(
            status_code=503,
            detail="No se pudo cargar el historial de ventas. Puede estar abierto en Excel.",
        )

    fecha_desde = _parse_iso_date(request.fecha_desde) if request.fecha_desde else None
    fecha_hasta = _parse_iso_date(request.fecha_hasta) if request.fecha_hasta else None

    resultados = _filtrar_historial(
        cache["filas"],
        busqueda=request.busqueda,
        cliente=request.cliente,
        codigo=request.codigo,
        moneda=request.moneda,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )

    output = _generar_excel_historial(resultados)
    filename = f"historial_ventas_{date.today().isoformat()}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/historial/metadata")
def historial_ventas_metadata(user: dict = Depends(get_current_user)):
    """Devuelve listas únicas de clientes y códigos del historial de ventas."""
    cache = _get_cached_historial()
    if not cache["filas"]:
        from app.config import get_settings
        from pathlib import Path

        settings = get_settings()
        excel_path = Path(settings.ventas_facturacion_excel_path)
        if not excel_path.exists():
            raise HTTPException(
                status_code=503,
                detail=f"No se encontró el archivo de ventas: {excel_path}",
            )
        raise HTTPException(
            status_code=503,
            detail="No se pudo cargar el historial de ventas. Puede estar abierto en Excel.",
        )

    return {"clientes": cache["clientes"], "codigos": cache["codigos"]}
