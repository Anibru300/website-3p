"""Pruebas del reporte exportable de analytics (Excel y PDF)."""

import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest

_TMP_DB = Path(__file__).parent / "_test_users.db"
os.environ["USERS_DB_PATH"] = str(_TMP_DB)
os.environ.setdefault("JWT_SECRET", "test-secret-de-pruebas-suficientemente-largo")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.analytics.reporte import (  # noqa: E402
    generar_reporte_excel,
    generar_reporte_pdf,
)
from app.analytics.router import datos_publicos  # noqa: E402
from app.database import users_connection  # noqa: E402

LOGO = str(Path(__file__).resolve().parent.parent.parent / "public/images/logo-3p-header.png")


@pytest.fixture()
def datos_con_contenido():
    ahora = datetime.now(timezone.utc)
    with users_connection() as conn:
        conn.execute("DELETE FROM analytics_events")
        # 3 eventos hoy, 1 ayer, todos públicos
        for i, delta in enumerate([0, 0, 0, 1]):
            ts = (ahora - timedelta(days=delta, hours=i)).strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                """
                INSERT INTO analytics_events
                (session_id, event_type, path, ip, device_type, browser, os, country, city, created_at)
                VALUES (?, 'page_view', '/', '187.190.10.20', 'desktop', 'Chrome', 'Windows', 'México', 'León', ?)
                """,
                (f"sess-{i}", ts),
            )
        # 1 evento de panel admin (debe EXCLUIRSE del reporte público)
        conn.execute(
            """
            INSERT INTO analytics_events
            (session_id, event_type, path, ip, created_at)
            VALUES ('sess-admin', 'section_view', '/admin', '127.0.0.1', ?)
            """,
            (ahora.strftime("%Y-%m-%d %H:%M:%S"),),
        )
        conn.commit()
    datos = datos_publicos(7)
    yield datos
    with users_connection() as conn:
        conn.execute("DELETE FROM analytics_events")
        conn.commit()


@pytest.fixture()
def datos_vacios():
    with users_connection() as conn:
        conn.execute("DELETE FROM analytics_events")
        conn.commit()
    return datos_publicos(7)


class TestDatosPublicos:
    def test_excluye_trafico_admin(self, datos_con_contenido):
        total = sum(d["total"] for d in datos_con_contenido["por_dia"])
        assert total == 4  # el evento /admin no cuenta

    def test_agrupa_paises(self, datos_con_contenido):
        paises = {p["nombre"]: p["total"] for p in datos_con_contenido["paises"]}
        assert paises.get("México") == 4


class TestReporteExcel:
    def test_excel_valido_con_datos(self, datos_con_contenido):
        contenido = generar_reporte_excel(datos_con_contenido, 7, None, None, logo_path=LOGO)
        assert contenido[:2] == b"PK"  # ZIP válido (xlsx)

        from openpyxl import load_workbook

        wb = load_workbook(__import__("io").BytesIO(contenido))
        assert "Resumen" in wb.sheetnames
        assert "Visitas" in wb.sheetnames
        ws = wb["Resumen"]
        # Logo incrustado
        assert len(ws._images) >= 1
        # Título presente
        celdas = [c.value for row in ws.iter_rows(min_row=1, max_row=15, max_col=1) for c in row]
        assert any(v and "Reporte de tráfico" in str(v) for v in celdas)

    def test_totales_coinciden_con_datos(self, datos_con_contenido):
        contenido = generar_reporte_excel(datos_con_contenido, 7, None, None, logo_path=None)
        from openpyxl import load_workbook

        wb = load_workbook(__import__("io").BytesIO(contenido))
        ws = wb["Resumen"]
        valores = {}
        for row in ws.iter_rows(min_row=1, max_col=2):
            if row[0].value and row[0].value not in valores:
                valores[row[0].value] = row[1].value
        assert valores["Total de eventos"] == 4
        assert valores["Países"] == 1

    def test_excel_sin_datos_no_rompe(self, datos_vacios):
        contenido = generar_reporte_excel(datos_vacios, 7, None, None, logo_path=None)
        assert contenido[:2] == b"PK"
        from openpyxl import load_workbook

        wb = load_workbook(__import__("io").BytesIO(contenido))
        assert "Resumen" in wb.sheetnames

    def test_etiqueta_periodo_fechas(self, datos_vacios):
        from app.analytics.reporte import _etiqueta_periodo

        assert _etiqueta_periodo(30, None, None) == "Últimos 30 días"
        assert _etiqueta_periodo(30, "2026-08-01", "2026-08-31") == "2026-08-01 a 2026-08-31"


class TestReportePdf:
    def test_pdf_valido_con_datos(self, datos_con_contenido):
        contenido = generar_reporte_pdf(datos_con_contenido, 7, None, None, logo_path=LOGO)
        assert contenido[:4] == b"%PDF"

        from PyPDF2 import PdfReader

        reader = PdfReader(__import__("io").BytesIO(contenido))
        assert len(reader.pages) >= 1
        texto = "".join(page.extract_text() or "" for page in reader.pages)
        assert "Reporte de tráfico" in texto
        assert "México" in texto

    def test_pdf_sin_datos_no_rompe(self, datos_vacios):
        contenido = generar_reporte_pdf(datos_vacios, 7, None, None, logo_path=None)
        assert contenido[:4] == b"%PDF"
        from PyPDF2 import PdfReader

        reader = PdfReader(__import__("io").BytesIO(contenido))
        assert len(reader.pages) >= 1


class TestEndpointsReporte:
    def test_endpoints_requieren_admin(self, datos_con_contenido):
        from fastapi.testclient import TestClient

        from app.main import app

        with TestClient(app) as c:
            r_excel = c.get("/api/analytics/reporte/excel?dias=7")
            r_pdf = c.get("/api/analytics/reporte/pdf?dias=7")
        assert r_excel.status_code in (401, 403, 422)
        assert r_pdf.status_code in (401, 403, 422)

    def test_endpoints_con_admin_devuelven_archivo(self, datos_con_contenido):
        from fastapi.testclient import TestClient

        from app.auth.dependencies import require_admin
        from app.main import app

        app.dependency_overrides[require_admin] = lambda: {"email": "admin@test", "role": "admin"}
        try:
            with TestClient(app) as c:
                r_excel = c.get("/api/analytics/reporte/excel?dias=7")
                r_pdf = c.get("/api/analytics/reporte/pdf?dias=7")
        finally:
            app.dependency_overrides.clear()

        assert r_excel.status_code == 200
        assert r_excel.content[:2] == b"PK"
        assert "attachment" in r_excel.headers.get("content-disposition", "")
        assert r_pdf.status_code == 200
        assert r_pdf.content[:4] == b"%PDF"
        assert "attachment" in r_pdf.headers.get("content-disposition", "")


def teardown_module():
    try:
        if _TMP_DB.exists():
            _TMP_DB.unlink()
    except OSError:
        pass
    for suffix in ("-wal", "-shm", "-journal"):
        p = Path(str(_TMP_DB) + suffix)
        if p.exists():
            try:
                p.unlink()
            except OSError:
                pass
