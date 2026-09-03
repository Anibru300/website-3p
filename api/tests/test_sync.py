"""Pruebas del sync Excel -> SQLite (Fase 2): roundtrip, feature flag y fallback."""

import os
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

_TMP_DB = Path(__file__).parent / "_test_users.db"
os.environ["USERS_DB_PATH"] = str(_TMP_DB)
os.environ.setdefault("JWT_SECRET", "test-secret-de-pruebas-suficientemente-largo")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.config import get_settings  # noqa: E402
from app.database import users_connection  # noqa: E402
from app.services.excel import (  # noqa: E402
    _fotos_from_rows,
    _material_en_vales_from_rows,
    _pedidos_vivos_from_rows,
    _vales_abiertos_from_rows,
    get_fotos_map,
    get_material_en_vales_by_code,
    get_pedidos_vivos_excel,
    get_vales_abiertos_count,
)
from app.sync import db as sync_db  # noqa: E402
from app.sync.job import sincronizar_fuente  # noqa: E402


def _crear_excel_vales(ruta: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "VALES"
    ws.append(["FOLIO_VALE", "STATUS", "ENTREGADO_A", "FECHA_SALIDA"])
    ws.append(["V-1", "ABIERTO", "JOAN PEREZ", "2026-09-01"])
    ws.append(["V-2", "ABIERTO", "ABELARDO RUIZ", "2026-09-02"])
    ws.append(["V-3", "CERRADO", "OTRO", "2026-08-30"])
    ws2 = wb.create_sheet("DETALLE_VALES")
    ws2.append(["FOLIO_VALE", "CODIGO", "DESCRIPCION", "CANTIDAD", "CANTIDAD_VIVA", "ALMACEN_ORIGEN", "STATUS"])
    ws2.append(["V-1", "A-001", "Motor", 2, 2, "A-1 ROXELL", "ABIERTO"])
    ws2.append(["V-1", "A-002", "Banda", 1, 1, "A-1 ROXELL", "ABIERTO"])
    ws2.append(["V-2", "A-001", "Motor", 3, 3, "A-20 SBM", "ABIERTO"])
    ws2.append(["V-2", "A-003", "Foco", 1, 0, "A-20 SBM", "CERRADO"])
    ws2.append(["V-3", "A-004", "Tornillo", 5, 0, "A-1 ROXELL", "CERRADO"])
    ws3 = wb.create_sheet("FOTOS_PRODUCTOS")
    ws3.append(["CODIGO", "RUTA_FOTO"])
    ws3.append(["A-001", "Y:/fotos/a1.jpg"])
    ws3.append(["A-002", "Y:/fotos/a2.jpg"])
    wb.save(ruta)
    wb.close()


def _crear_excel_pedidos(ruta: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "PEDIDOS"
    ws.append(["FOLIO_PEDIDO", "FECHA_PEDIDO", "CLIENTE", "STATUS_GENERAL", "MONEDA"])
    ws.append(["P-10", "2026-09-01", "GRANJA SAN JUAN", "ACTIVO", "MXN"])
    ws.append(["P-11", "2026-09-02", "AVICOLA DEL NORTE", "ACTIVO", "USD"])
    ws2 = wb.create_sheet("DETALLE_PEDIDOS")
    ws2.append([
        "FOLIO_PEDIDO", "FECHA_PEDIDO", "CLIENTE", "CODIGO", "DESCRIPCION",
        "CANT_PEDIDA", "PRECIO_UNITARIO", "TOTAL_FACTURADO", "PENDIENTE_FACTURAR", "MONEDA",
    ])
    ws2.append(["P-10", "2026-09-01", "GRANJA SAN JUAN", "A-001", "Motor", 2, 100.0, 0, 200.0, "MXN"])
    ws2.append(["P-11", "2026-09-02", "AVICOLA DEL NORTE", "A-002", "Banda", 1, 50.0, 50.0, 0, "USD"])
    wb.save(ruta)
    wb.close()


@pytest.fixture()
def excel_vales(tmp_path, monkeypatch):
    ruta = tmp_path / "vales.xlsx"
    _crear_excel_vales(ruta)
    monkeypatch.setenv("VALES_EXCEL_PATH", str(ruta))
    get_settings.cache_clear()
    yield ruta
    get_settings.cache_clear()


@pytest.fixture()
def excel_pedidos(tmp_path, monkeypatch):
    ruta = tmp_path / "pedidos.xlsx"
    _crear_excel_pedidos(ruta)
    monkeypatch.setenv("PEDIDOS_PENDIENTES_FACTURAR_EXCEL_PATH", str(ruta))
    get_settings.cache_clear()
    yield ruta
    get_settings.cache_clear()


@pytest.fixture(autouse=True)
def limpiar_sync():
    with users_connection() as conn:
        sync_db.init_sync_db(conn)
        conn.execute("DELETE FROM sync_sheets")
        conn.execute("DELETE FROM sync_log")
        conn.commit()
    yield
    with users_connection() as conn:
        conn.execute("DELETE FROM sync_sheets")
        conn.execute("DELETE FROM sync_log")
        conn.commit()


class TestSyncBasico:
    def test_sync_ok_guarda_hojas_y_log(self, excel_vales):
        resumen = sincronizar_fuente("vales")
        assert resumen["estado"] == "ok"
        assert resumen["filas"] > 0
        with users_connection() as conn:
            ultimo = sync_db.ultimo_sync(conn, "vales")
            assert ultimo["estado"] == "ok"
            sheets = sync_db.cargar_sheets(conn, "vales")
        assert set(sheets.keys()) == {"VALES", "DETALLE_VALES", "FOTOS_PRODUCTOS"}
        assert len(sheets["DETALLE_VALES"]) == 5

    def test_sync_fuente_no_habilitada(self, excel_vales):
        with pytest.raises(ValueError):
            sincronizar_fuente("ventas_facturacion")

    def test_sync_con_hoja_renombrada_registra_error(self, excel_vales):
        from openpyxl import load_workbook

        wb = load_workbook(excel_vales)
        wb["DETALLE_VALES"].title = "DETALLE"
        wb.save(excel_vales)
        wb.close()

        resumen = sincronizar_fuente("vales")
        assert resumen["estado"] == "error"
        assert "DETALLE_VALES" in resumen["error"]
        with users_connection() as conn:
            ultimo = sync_db.ultimo_sync(conn, "vales")
            assert ultimo["estado"] == "error"


class TestFeatureFlag:
    def _activar_sync(self, monkeypatch):
        monkeypatch.setenv("USE_SYNC_TABLES", "true")
        get_settings.cache_clear()

    def test_flag_off_usa_excel_en_vivo(self, excel_vales, monkeypatch):
        sincronizar_fuente("vales")
        monkeypatch.setenv("USE_SYNC_TABLES", "false")
        get_settings.cache_clear()
        assert get_material_en_vales_by_code() == {"A-001": 5.0, "A-002": 1.0}
        assert get_vales_abiertos_count() == 2

    def test_flag_on_datos_identicos_al_excel(self, excel_vales, monkeypatch):
        # Datos vía Excel en vivo (núcleos con rows del live loader)
        from app.services.excel import _load_fotos_rows_excel, _load_vales_excel

        cabs_live, dets_live = _load_vales_excel()
        material_live = _material_en_vales_from_rows(cabs_live, dets_live)
        abiertos_live = _vales_abiertos_from_rows(cabs_live, dets_live)
        fotos_live = _fotos_from_rows(_load_fotos_rows_excel())

        sincronizar_fuente("vales")
        self._activar_sync(monkeypatch)

        assert get_material_en_vales_by_code() == material_live
        assert get_vales_abiertos_count() == abiertos_live

        # Limpiar caché de fotos para forzar re-lectura vía sync
        import app.services.excel as excel_mod

        excel_mod._FOTOS_CACHE = {"map": None, "ts": 0}
        assert get_fotos_map() == fotos_live

    def test_pedidos_identicos_ambas_rutas(self, excel_pedidos, monkeypatch):
        from app.services.excel import _load_pedidos_excel

        cabs, dets = _load_pedidos_excel()
        esperado = _pedidos_vivos_from_rows(cabs, dets)

        sincronizar_fuente("pedidos")
        self._activar_sync(monkeypatch)

        assert get_pedidos_vivos_excel() == esperado

    def test_sync_vacio_falla_a_excel_en_vivo(self, excel_vales, monkeypatch):
        """Sin sync previo, aunque el flag esté activo, el getter lee el Excel."""
        self._activar_sync(monkeypatch)
        assert get_material_en_vales_by_code() == {"A-001": 5.0, "A-002": 1.0}


def teardown_module():
    try:
        if _TMP_DB.exists():
            _TMP_DB.unlink()
    except OSError:
        pass
    for suffix in ("-wal", "-shm", "-journal"):
        p = Path(str(_TMP_DB) + suffix)
        if p.exists():
            try:
                p.unlink()
            except OSError:
                pass
