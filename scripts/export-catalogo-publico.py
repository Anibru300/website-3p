#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Exporta el catálogo público de marcas para el sitio web 3P.

Lee (SOLO LECTURA) la hoja FOTOS_PRODUCTOS y las hojas de inventario por marca
de BD_ALMACEN_3P.xlsx, optimiza las fotos con Pillow (máx. 1200 px de lado
mayor, RGB, JPEG calidad 85) y las copia a public/images/catalogo/<slug>/.

Genera src/data/catalogoMarcas.js con la estructura consumida por
GenericBrandPage.jsx (export const brandCatalogs).

Marcas exportadas al sitio público: fancom, lubing, sbm, lbwhite,
georgia-poultry, amt, alke.

Marcas EXCLUIDAS a propósito (decisión de negocio):
  - chore-time    : 3P no es distribuidor autorizado, no se publica.
  - roxell        : sin tarjeta/logo en el sitio (se reporta aparte).
  - landmeco      : sin tarjeta/logo en el sitio (se reporta aparte).
  - ms-detergentes: ese catálogo ya existe hecho a mano en
                    src/data/msSchippersData.js; no se regenera.

El script es idempotente: borra y recrea las carpetas de las marcas que
exporta. NO toca las carpetas ms-detergentes ni chore-time de
public/images, y nunca escribe en el Excel.

Uso (desde la raíz del proyecto):
    api/.venv/Scripts/python scripts/export-catalogo-publico.py
"""

import re
import sys
from pathlib import Path

import openpyxl
from PIL import Image

# --- Configuración ---------------------------------------------------------

EXCEL_PATH = Path(r'Y:/ALMACEN/Mejora Continua ALMACEN/Nuevo Control de Almacen/BASE DE DATOS/BD_ALMACEN_3P.xlsx')
PROJECT_ROOT = Path(__file__).resolve().parent.parent
PUBLIC_CATALOGO = PROJECT_ROOT / 'public' / 'images' / 'catalogo'
OUTPUT_JS = PROJECT_ROOT / 'src' / 'data' / 'catalogoMarcas.js'

MAX_SIDE = 1200
JPEG_QUALITY = 85

# Carpeta de foto en disco (tal cual aparece en RUTA_FOTO) -> slug del sitio
FOLDER_TO_SLUG = {
    '40 A-40 FANCOM': 'fancom',
    '2 A-2 LUBING': 'lubing',
    '20 A-20 SBM': 'sbm',
    '26 A-26 L.B. WHITE': 'lbwhite',
    '10 A-10 GEORGIA POULTRY': 'georgia-poultry',
    '45 A-45 AMT INC': 'amt',
    '48 A-48 ALKE': 'alke',
}

# Hoja de inventario de marca -> slug del sitio
SHEET_TO_SLUG = {
    '40 A-40 FANCOM': 'fancom',
    '2 A-2 LUBING': 'lubing',
    '20 A-20 SBM': 'sbm',
    '26 A-26 L.B. WHITE': 'lbwhite',
    '10 A-10 GEORGIA POULTRY': 'georgia-poultry',
    '45 A-45 AMT INC': 'amt',
    '48 A-48 ALKE': 'alke',
}

# Slug -> nombre visible en el sitio
SLUG_TO_NAME = {
    'fancom': 'FANCOM',
    'lubing': 'LUBING',
    'sbm': 'SBM',
    'lbwhite': 'L.B. White',
    'georgia-poultry': 'Georgia Poultry',
    'amt': 'AMT',
    'alke': 'ALKE',
}


def js_str(valor):
    """Escapa un string para usarlo como literal JSON/JS con comillas dobles."""
    return '"' + str(valor).replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n') + '"'


def slugify_codigo(codigo):
    """Slug de un código de producto: minúsculas, espacios -> guion,
    sin caracteres inválidos para URL/Windows."""
    slug = str(codigo).strip().lower()
    slug = slug.replace(' ', '-')
    slug = re.sub(r'[^a-z0-9._-]', '', slug)
    slug = slug.strip('.-_') or 'sin-codigo'
    return slug


def cargar_descripciones(wb):
    """Lee las hojas de marca y devuelve {slug: {codigo: descripcion}}."""
    result = {}
    for hoja, slug in SHEET_TO_SLUG.items():
        if hoja not in wb.sheetnames:
            print(f'  !! Hoja no encontrada: {hoja}')
            continue
        ws = wb[hoja]
        mapping = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            codigo = row[0]
            descripcion = row[1]
            if codigo is None:
                continue
            mapping[str(codigo).strip()] = str(descripcion).strip() if descripcion else ''
        result[slug] = mapping
    return result


def optimizar_foto(src, dst):
    """Abre la foto, la convierte a RGB, la redimensiona (máx MAX_SIDE)
    y la guarda como JPEG calidad 85. Devuelve True si tuvo éxito."""
    try:
        with Image.open(src) as img:
            img = img.convert('RGB')
            if max(img.size) > MAX_SIDE:
                img.thumbnail((MAX_SIDE, MAX_SIDE), Image.LANCZOS)
            img.save(dst, 'JPEG', quality=JPEG_QUALITY, optimize=True)
        return True
    except Exception as exc:
        print(f'  !! Error procesando {src}: {exc}')
        return False


def main():
    print(f'Leyendo Excel (solo lectura): {EXCEL_PATH}')
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    descripciones = cargar_descripciones(wb)
    print(f'Hojas de marca cargadas: {len(descripciones)}')

    ws = wb['FOTOS_PRODUCTOS']
    fotos = []  # (slug, codigo, ruta_foto)
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo, ruta_foto = row[0], row[1]
        if not codigo or not ruta_foto:
            continue
        parts = str(ruta_foto).rsplit('\\', 2)
        if len(parts) < 3:
            print(f'  !! Ruta inesperada: {ruta_foto}')
            continue
        folder = parts[-2]
        slug = FOLDER_TO_SLUG.get(folder)
        if not slug:
            continue  # marca excluida (roxell, landmeco, chore-time, ms-detergentes)
        fotos.append((slug, str(codigo).strip(), Path(str(ruta_foto))))

    print(f'Fotos en marcas exportables: {len(fotos)}')

    # Idempotente: borrar y recrear las carpetas de las marcas que exporta
    for slug in FOLDER_TO_SLUG.values():
        marca_dir = PUBLIC_CATALOGO / slug
        if marca_dir.exists():
            for f in marca_dir.iterdir():
                if f.is_file():
                    f.unlink()
        marca_dir.mkdir(parents=True, exist_ok=True)

    # Copiar y optimizar fotos
    sin_descripcion = {}
    foto_faltante = {}
    exportadas = {}

    for slug, codigo, src in fotos:
        desc_map = descripciones.get(slug, {})
        descripcion = desc_map.get(codigo, '')
        if not descripcion:
            sin_descripcion.setdefault(slug, []).append(codigo)

        archivo = slugify_codigo(codigo) + '.jpg'
        dst = PUBLIC_CATALOGO / slug / archivo

        if not src.exists():
            foto_faltante.setdefault(slug, []).append(codigo)
            continue
        if not optimizar_foto(src, dst):
            foto_faltante.setdefault(slug, []).append(codigo)
            continue

        exportadas.setdefault(slug, []).append({
            'codigo': codigo,
            'descripcion': descripcion,
            'imagen': f'/images/catalogo/{slug}/{archivo}',
        })

    # Generar src/data/catalogoMarcas.js
    lineas = []
    lineas.append('// GENERADO AUTOMÁTICAMENTE por scripts/export-catalogo-publico.py')
    lineas.append('// No editar a mano: se regenera al volver a correr el script.')
    lineas.append('// Fuente: BD_ALMACEN_3P.xlsx (hoja FOTOS_PRODUCTOS + hojas de marca).')
    lineas.append('')
    lineas.append('export const brandCatalogs = {')

    marcas_con_datos = 0
    for slug in sorted(exportadas):
        productos = sorted(exportadas[slug], key=lambda p: p['codigo'].lower())
        if not productos:
            continue
        marcas_con_datos += 1
        lineas.append(f'  {js_str(slug)}: {{')
        lineas.append(f'    nombre: {js_str(SLUG_TO_NAME[slug])},')
        lineas.append('    productos: [')
        for p in productos:
            lineas.append('      {')
            lineas.append(f'        codigo: {js_str(p["codigo"])},')
            lineas.append(f'        descripcion: {js_str(p["descripcion"])},')
            lineas.append(f'        imagen: {js_str(p["imagen"])},')
            lineas.append('      },')
        lineas.append('    ],')
        lineas.append('  },')

    lineas.append('};')
    lineas.append('')
    OUTPUT_JS.write_text('\n'.join(lineas), encoding='utf-8')

    # Resumen
    print()
    print('=== RESUMEN ===')
    total = 0
    for slug in sorted(exportadas):
        n = len(exportadas[slug])
        total += n
        print(f'  {slug}: {n} fotos exportadas')
    print(f'  TOTAL exportadas: {total}')
    print(f'  Marcas en catalogoMarcas.js: {marcas_con_datos}')
    if sin_descripcion:
        print('  Codigos SIN descripcion en la hoja de marca:')
        for slug, codigos in sorted(sin_descripcion.items()):
            print(f'    {slug}: {codigos}')
    else:
        print('  Codigos sin descripcion: ninguno')
    if foto_faltante:
        print('  Codigos con foto FALTANTE / error al procesar:')
        for slug, codigos in sorted(foto_faltante.items()):
            print(f'    {slug}: {codigos}')
    else:
        print('  Fotos faltantes: ninguna')
    print(f'  Archivo generado: {OUTPUT_JS}')

    return 0


if __name__ == '__main__':
    sys.exit(main())
